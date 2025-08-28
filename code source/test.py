#
#                  V 1.07
#               Inzerty by RLK
#

import os
import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets, uic
from PyQt5 import sip
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QProgressBar
import sys
import openpyxl
from openpyxl import load_workbook
import logging
logging.basicConfig(filename='Controle-Paie-LOGS.log',level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
import numpy as np
from gestion_des_agences import CORRESPONDANCES_AGENCES
import gestion_des_agences

# Traitement des Images: 
from os import path
path_to_LogoJanus = path.abspath(path.join(path.dirname(__file__), 'LogoJanus.png'))
path_to_LogoJanus = path_to_LogoJanus.replace("\\", "\\\\")

path_to_LogoAppli = path.abspath(path.join(path.dirname(__file__), 'pngegg0.png'))
path_to_LogoAppli = path_to_LogoAppli.replace("\\", "\\\\")

path_to_check_mark = path.abspath(path.join(path.dirname(__file__), 'check-mark-png-45026-resize.png'))
path_to_check_mark = path_to_check_mark.replace("\\", "\\\\")


###########################################
#             gestion_paie_ui             #
###########################################

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(772, 631)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.setObjectName("verticalLayout")
        self.frame_2_logo = QtWidgets.QFrame(self.centralwidget)
        self.frame_2_logo.setMinimumSize(QtCore.QSize(450, 200))
        self.frame_2_logo.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_2_logo.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2_logo.setObjectName("frame_2_logo")
        self.label_2 = QtWidgets.QLabel(self.frame_2_logo)
        self.label_2.setGeometry(QtCore.QRect(10, 20, 431, 171))
        self.label_2.setMinimumSize(QtCore.QSize(0, 0))
        font = QtGui.QFont()
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setText("")
        
        self.label_2.setPixmap(QtGui.QPixmap(path_to_LogoJanus))
        
        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.frame_2_logo, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.frame_1_text = QtWidgets.QFrame(self.centralwidget)
        self.frame_1_text.setMinimumSize(QtCore.QSize(450, 250))
        self.frame_1_text.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_1_text.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_1_text.setObjectName("frame_1_text")
        self.Bienvenue = QtWidgets.QLabel(self.frame_1_text)
        self.Bienvenue.setGeometry(QtCore.QRect(90, 120, 271, 41))
        font = QtGui.QFont()
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.Bienvenue.setFont(font)
        self.Bienvenue.setLocale(QtCore.QLocale(QtCore.QLocale.French, QtCore.QLocale.France))
        self.Bienvenue.setObjectName("Bienvenue")
        self.versiondedev = QtWidgets.QLabel(self.frame_1_text)
        self.versiondedev.setGeometry(QtCore.QRect(110, 160, 241, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.versiondedev.setFont(font)
        self.versiondedev.setLocale(QtCore.QLocale(QtCore.QLocale.French, QtCore.QLocale.France))
        self.versiondedev.setObjectName("versiondedev")
        self.Bienvenue_3 = QtWidgets.QLabel(self.frame_1_text)
        self.Bienvenue_3.setGeometry(QtCore.QRect(110, 200, 241, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.Bienvenue_3.setFont(font)
        self.Bienvenue_3.setLocale(QtCore.QLocale(QtCore.QLocale.French, QtCore.QLocale.France))
        self.Bienvenue_3.setObjectName("Bienvenue_3")
        self.logoexcel = QtWidgets.QLabel(self.frame_1_text)
        self.logoexcel.setGeometry(QtCore.QRect(120, 10, 231, 111))
        self.logoexcel.setText("")

        self.logoexcel.setPixmap(QtGui.QPixmap(path_to_LogoAppli))

        self.logoexcel.setObjectName("logoexcel")
        self.verticalLayout.addWidget(self.frame_1_text, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.frame_bouton = QtWidgets.QFrame(self.centralwidget)
        self.frame_bouton.setMinimumSize(QtCore.QSize(450, 100))
        self.frame_bouton.setLocale(QtCore.QLocale(QtCore.QLocale.French, QtCore.QLocale.France))
        self.frame_bouton.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_bouton.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_bouton.setObjectName("frame_bouton")
        self.pushButton_Annuler = QtWidgets.QPushButton(self.frame_bouton)
        self.pushButton_Annuler.setGeometry(QtCore.QRect(80, 0, 121, 41))
        self.pushButton_Annuler.setLocale(QtCore.QLocale(QtCore.QLocale.French, QtCore.QLocale.France))
        self.pushButton_Annuler.setAutoDefault(True)
        self.pushButton_Annuler.setDefault(True)
        self.pushButton_Annuler.setFlat(False)
        self.pushButton_Annuler.setObjectName("pushButton_Annuler")
        self.pushButton_Continuer = QtWidgets.QPushButton(self.frame_bouton)
        self.pushButton_Continuer.setGeometry(QtCore.QRect(260, 0, 121, 41))
        self.pushButton_Continuer.setLocale(QtCore.QLocale(QtCore.QLocale.French, QtCore.QLocale.France))
        self.pushButton_Continuer.setAutoDefault(False)
        self.pushButton_Continuer.setDefault(False)
        self.pushButton_Continuer.setFlat(False)
        self.pushButton_Continuer.setObjectName("pushButton_Continuer")
        self.version = QtWidgets.QLabel(self.frame_bouton)
        self.version.setGeometry(QtCore.QRect(140, 70, 191, 22))
        self.version.setObjectName("version")
        self.verticalLayout.addWidget(self.frame_bouton, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Contrôle Paie"))
        self.Bienvenue.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:14pt;\">Contrôle Paie</span></p></body></html>"))
        self.versiondedev.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:400;\">Bienvenue</span></p></body></html>"))
        self.Bienvenue_3.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:400;\">Souhaitez vous continuer ?</span></p></body></html>"))
        self.pushButton_Annuler.setText(_translate("MainWindow", "Annuler"))
        self.pushButton_Continuer.setText(_translate("MainWindow", "Continuer"))
        self.version.setText(_translate("MainWindow", "Version 1.07-1 05/2024  © Inzerty RLK "))


###########################################
#           gestion_paie_ui_dialog        #
###########################################

class Ui_Dialog(object):
    def setupUi1(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(660, 636)
        Dialog.setSizeGripEnabled(False)
        Dialog.setModal(False)
        self.verticalLayout = QtWidgets.QVBoxLayout(Dialog)
        self.verticalLayout.setObjectName("verticalLayout")
        self.logo = QtWidgets.QFrame(Dialog)
        self.logo.setMinimumSize(QtCore.QSize(600, 150))
        self.logo.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.logo.setFrameShadow(QtWidgets.QFrame.Raised)
        self.logo.setObjectName("logo")
        self.label_2 = QtWidgets.QLabel(self.logo)
        self.label_2.setGeometry(QtCore.QRect(180, 10, 231, 141))
        self.label_2.setMinimumSize(QtCore.QSize(0, 0))
        font = QtGui.QFont()
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setText("")

        self.label_2.setPixmap(QtGui.QPixmap(path_to_LogoAppli))

        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.logo, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.Choix_agence = QtWidgets.QFrame(Dialog)
        self.Choix_agence.setMinimumSize(QtCore.QSize(600, 100))
        self.Choix_agence.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.Choix_agence.setFrameShadow(QtWidgets.QFrame.Raised)
        self.Choix_agence.setObjectName("Choix_agence")
        self.comboBox = QtWidgets.QComboBox(self.Choix_agence)
        self.comboBox.setGeometry(QtCore.QRect(40, 50, 301, 31))
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItems(CORRESPONDANCES_AGENCES.keys())

        self.pushButton = QtWidgets.QPushButton(self.Choix_agence)
        self.pushButton.setGeometry(QtCore.QRect(390, 50, 181, 31))
        self.pushButton.setObjectName("pushButton")
        self.label = QtWidgets.QLabel(self.Choix_agence)
        self.label.setGeometry(QtCore.QRect(50, 10, 291, 22))
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.Choix_agence, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.choix_fichiers = QtWidgets.QFrame(Dialog)
        self.choix_fichiers.setMinimumSize(QtCore.QSize(600, 250))
        self.choix_fichiers.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.choix_fichiers.setFrameShadow(QtWidgets.QFrame.Raised)
        self.choix_fichiers.setObjectName("choix_fichiers")
        self.First_excel = QtWidgets.QLabel(self.choix_fichiers)
        self.First_excel.setGeometry(QtCore.QRect(60, 10, 441, 22))
        self.First_excel.setObjectName("First_excel")
        self.First_excel_2 = QtWidgets.QLabel(self.choix_fichiers)
        self.First_excel_2.setGeometry(QtCore.QRect(60, 60, 441, 22))
        self.First_excel_2.setObjectName("First_excel_2")
        self.First_excel_3 = QtWidgets.QLabel(self.choix_fichiers)
        self.First_excel_3.setGeometry(QtCore.QRect(60, 110, 441, 22))
        self.First_excel_3.setObjectName("First_excel_3")
        self.First_excel_4 = QtWidgets.QLabel(self.choix_fichiers)
        self.First_excel_4.setGeometry(QtCore.QRect(60, 170, 441, 22))
        self.First_excel_4.setObjectName("First_excel_4")
        self.pushButton_Cotisation = QtWidgets.QPushButton(self.choix_fichiers)
        self.pushButton_Cotisation.setGeometry(QtCore.QRect(430, 10, 61, 31))
        self.pushButton_Cotisation.setObjectName("pushButton_Cotisation")
        self.pushButton_Matricule = QtWidgets.QPushButton(self.choix_fichiers)
        self.pushButton_Matricule.setGeometry(QtCore.QRect(430, 60, 61, 31))
        self.pushButton_Matricule.setObjectName("pushButton_Matricule")
        self.pushButton_Rubrique = QtWidgets.QPushButton(self.choix_fichiers)
        self.pushButton_Rubrique.setGeometry(QtCore.QRect(430, 100, 61, 31))
        self.pushButton_Rubrique.setObjectName("pushButton_Rubrique")
        self.pushButton_Sortie = QtWidgets.QPushButton(self.choix_fichiers)
        self.pushButton_Sortie.setGeometry(QtCore.QRect(430, 160, 61, 31))
        self.pushButton_Sortie.setObjectName("pushButton_Sortie")
        self.IF_OK_excel = QtWidgets.QLabel(self.choix_fichiers)
        self.IF_OK_excel.setGeometry(QtCore.QRect(520, 10, 31, 31))
        self.IF_OK_excel.setInputMethodHints(QtCore.Qt.ImhNone)
        self.IF_OK_excel.setText("")

        self.IF_OK_excel.setPixmap(QtGui.QPixmap(path_to_check_mark))
        self.IF_OK_excel.setScaledContents(True)
        self.IF_OK_excel.setObjectName("IF_OK_excel")
        self.IF_OK_excel.setVisible(False)

        self.IF_OK_excel_2 = QtWidgets.QLabel(self.choix_fichiers)
        self.IF_OK_excel_2.setGeometry(QtCore.QRect(520, 50, 31, 31))
        self.IF_OK_excel_2.setText("")

        self.IF_OK_excel_2.setPixmap(QtGui.QPixmap(path_to_check_mark))
        self.IF_OK_excel_2.setScaledContents(True)
        self.IF_OK_excel_2.setObjectName("IF_OK_excel_2")
        self.IF_OK_excel_2.setVisible(False)

        self.IF_OK_excel_3 = QtWidgets.QLabel(self.choix_fichiers)
        self.IF_OK_excel_3.setGeometry(QtCore.QRect(520, 90, 31, 31))
        self.IF_OK_excel_3.setText("")

        self.IF_OK_excel_3.setPixmap(QtGui.QPixmap(path_to_check_mark))
        self.IF_OK_excel_3.setScaledContents(True)
        self.IF_OK_excel_3.setObjectName("IF_OK_excel_3")
        self.IF_OK_excel_3.setVisible(False)

        self.IF_OK_excel_4 = QtWidgets.QLabel(self.choix_fichiers)
        self.IF_OK_excel_4.setGeometry(QtCore.QRect(520, 150, 31, 31))
        self.IF_OK_excel_4.setText("")

        self.IF_OK_excel_4.setPixmap(QtGui.QPixmap(path_to_check_mark))
        self.IF_OK_excel_4.setScaledContents(True)
        self.IF_OK_excel_4.setObjectName("IF_OK_excel_4")
        self.IF_OK_excel_4.setVisible(False)

        self.verticalLayout.addWidget(self.choix_fichiers, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.frame = QtWidgets.QFrame(Dialog)
        self.frame.setMinimumSize(QtCore.QSize(600, 50))
        self.frame.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.pushButton_2 = QtWidgets.QPushButton(self.frame)
        self.pushButton_2.setGeometry(QtCore.QRect(80, 10, 121, 31))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QtWidgets.QPushButton(self.frame)
        self.pushButton_3.setGeometry(QtCore.QRect(380, 10, 121, 31))
        self.pushButton_3.setLocale(QtCore.QLocale(QtCore.QLocale.French, QtCore.QLocale.France))
        self.pushButton_3.setDefault(True)
        self.pushButton_3.setObjectName("pushButton_3")
        self.verticalLayout.addWidget(self.frame, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.pushButton.setText(_translate("Dialog", "Gestion des agences"))
        self.label.setText(_translate("Dialog", "Veuillez choisir votre agence :"))
        self.First_excel.setText(_translate("Dialog", "Sélectionnez le fichier Excel \'Cotisation\'"))
        self.First_excel_2.setText(_translate("Dialog", "Sélectionnez le fichier Excel \'Matricule\'"))
        self.First_excel_3.setText(_translate("Dialog", "Sélectionnez le fichier Excel \'Rubrique\'"))
        self.First_excel_4.setText(_translate("Dialog", "Choisissez le fichier Excel à écrire"))
        self.pushButton_Cotisation.setText(_translate("Dialog", "..."))
        self.pushButton_Matricule.setText(_translate("Dialog", "..."))
        self.pushButton_Rubrique.setText(_translate("Dialog", "..."))
        self.pushButton_Sortie.setText(_translate("Dialog", "..."))
        self.pushButton_2.setText(_translate("Dialog", "Quitter"))
        self.pushButton_3.setText(_translate("Dialog", "->Executer <-"))


# Définir la fonction pour mettre à jour les éléments de la ComboBox
def update_combobox_items(dialog_ui, agence):
    if agence in CORRESPONDANCES_AGENCES:
        dialog_ui.comboBox.clear()
        dialog_ui.comboBox.addItems(CORRESPONDANCES_AGENCES[agence].keys())

class ErrorDialog(QtWidgets.QDialog):
    def __init__(self, error_message):
        super(ErrorDialog, self).__init__()
        self.setWindowTitle("Erreur")
        
        # Label pour afficher le message d'erreur
        self.error_label = QtWidgets.QLabel(error_message)
        
        # Créer un QFrame pour encadrer le logo
        self.logo_frame = QtWidgets.QFrame()
        self.logo_frame.setFrameStyle(QtWidgets.QFrame.Box | QtWidgets.QFrame.Plain)  # Style du cadre
        self.logo_frame.setLineWidth(0)  # Épaisseur de la ligne du cadre
        
        # QLabel pour afficher le logo
        self.logo_label = QtWidgets.QLabel()
        logo_pixmap = QtGui.QPixmap(path_to_LogoAppli)
        self.logo_label.setPixmap(logo_pixmap)
        
        # Ajouter le logo_label au logo_frame
        logo_layout = QtWidgets.QHBoxLayout(self.logo_frame)
        logo_layout.addWidget(self.logo_label)
        logo_layout.setAlignment(QtCore.Qt.AlignCenter)  # Centrer horizontalement le contenu
        
        # Bouton OK
        self.ok_button = QtWidgets.QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        
        # Layout
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.error_label)
        layout.addWidget(self.logo_frame)  # Ajouter le cadre du logo
        layout.addWidget(self.ok_button)
        
        self.setLayout(layout)

class ErrorDialogAlreadyRunning(QtWidgets.QDialog):
    def __init__(self, error_message):
        super(ErrorDialogAlreadyRunning, self).__init__()
        self.setWindowTitle("Une instance du logiciel est déjà en cours d'exécution")
        
        self.error_label = QtWidgets.QLabel(error_message)
        
        # Créer un QFrame pour encadrer le logo
        self.logo_frame = QtWidgets.QFrame()
        self.logo_frame.setFrameStyle(QtWidgets.QFrame.Box | QtWidgets.QFrame.Plain)  # Style du cadre
        self.logo_frame.setLineWidth(0)  # Épaisseur de la ligne du cadre
        
        # QLabel pour afficher le logo
        self.logo_label = QtWidgets.QLabel()
        logo_pixmap = QtGui.QPixmap(path_to_LogoAppli)
        self.logo_label.setPixmap(logo_pixmap)
        
        # Ajouter le logo_label au logo_frame
        logo_layout = QtWidgets.QHBoxLayout(self.logo_frame)
        logo_layout.addWidget(self.logo_label)
        logo_layout.setAlignment(QtCore.Qt.AlignCenter)  # Centrer horizontalement le contenu
        
        self.test_button = QtWidgets.QPushButton("Quitter")
        self.test_button.clicked.connect(self.accept)
        
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.error_label)
        layout.addWidget(self.logo_frame)  # Ajouter le cadre du logo
        layout.addWidget(self.test_button)
        
        self.setLayout(layout)


class SuccessDialog(QtWidgets.QDialog):
    def __init__(self, message):
        super(SuccessDialog, self).__init__()
        self.setWindowTitle("Succès")
        
        self.success_label = QtWidgets.QLabel(message)
        
        # Créer un QFrame pour encadrer le logo
        self.logo_frame = QtWidgets.QFrame()
        self.logo_frame.setFrameStyle(QtWidgets.QFrame.Box | QtWidgets.QFrame.Plain)  # Style du cadre
        self.logo_frame.setLineWidth(0)  # Épaisseur de la ligne du cadre
        
        # QLabel pour afficher le logo
        self.logo_label = QtWidgets.QLabel()
        logo_pixmap = QtGui.QPixmap(path_to_LogoAppli)
        self.logo_label.setPixmap(logo_pixmap)
        
        # Ajouter le logo_label au logo_frame
        logo_layout = QtWidgets.QHBoxLayout(self.logo_frame)
        logo_layout.addWidget(self.logo_label)
        logo_layout.setAlignment(QtCore.Qt.AlignCenter)  # Centrer horizontalement le contenu
        
        self.ok_button = QtWidgets.QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.success_label)
        layout.addWidget(self.logo_frame)  # Ajouter le cadre du logo
        layout.addWidget(self.ok_button)
        
        self.setLayout(layout)

class OperationProgressDialog(QtWidgets.QDialog):
    def __init__(self):
        super(OperationProgressDialog, self).__init__()
        self.setWindowTitle("Opération en cours")
        
        # Message d'opération en cours
        operation_message = "Opération sur les fichiers en cours. Veuillez patienter."
        self.operation_label = QtWidgets.QLabel(operation_message)
        
        # Créer un QFrame pour encadrer le logo
        self.logo_frame = QtWidgets.QFrame()
        self.logo_frame.setFrameStyle(QtWidgets.QFrame.Box | QtWidgets.QFrame.Plain)  # Style du cadre
        self.logo_frame.setLineWidth(0)  # Épaisseur de la ligne du cadre
        
        # QLabel pour afficher le logo
        self.logo_label = QtWidgets.QLabel()
        logo_pixmap = QtGui.QPixmap(path_to_LogoAppli)
        self.logo_label.setPixmap(logo_pixmap)
        
        # Ajouter le logo_label au logo_frame
        logo_layout = QtWidgets.QHBoxLayout(self.logo_frame)
        logo_layout.addWidget(self.logo_label)
        logo_layout.setAlignment(QtCore.Qt.AlignCenter)  # Centrer horizontalement le contenu
        
        # Bouton OK
        self.ok_button = QtWidgets.QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.operation_label)
        layout.addWidget(self.logo_frame)  # Ajouter le cadre du logo
        layout.addWidget(self.ok_button)
        
        self.setLayout(layout)



class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.dialog = QtWidgets.QDialog()
        self.setupUi(self)
        self.dialog_ui = Ui_Dialog()
        self.dialog_ui.setupUi1(self.dialog)
        self.dialog_ui.comboBox.currentIndexChanged.connect(self.on_index_changed)
        self.pushButton_Annuler.clicked.connect(self.annuler)
        self.pushButton_Continuer.clicked.connect(self.continuer)
        


        # Déclaration des variables pour les fichiers Excel
        self.fileName_Cotisation = None
        self.fileName_Matricule = None
        self.fileName_Rubrique = None

        # Déclaration des variables necessaires:
        self.lignes_suivantes_colonne1_base_filtre = None
        self.correspondances_selectionnees = None
        self.rubriques_en_doublon_selectionnees = None

    ######################################################
    # fonctions des boutons du fichier gestion_paie_ui : #
    ######################################################
    def annuler(self):
         # Supprimer le fichier 'pid_projet_janus.tmp' s'il existe
        filename = "pid_projet_janus.tmp"
        if os.path.exists(filename):
            os.remove(filename)
        
        # Quitter l'application :
        print("Annuler button clicked")
        logging.info("Annuler button clicked")
        QtWidgets.QApplication.quit()

    def continuer(self):
        # Chemin vers le fichier
        filename = "pid_projet_janus.tmp"

        # Vérifier si le fichier existe
        if os.path.exists(filename):
            # Utiliser la fonction show_error_dialog pour afficher un message d'erreur
            self.show_error_running_dialog("Une instance du logiciel est déjà en cours d'exécution. Une seule instance du logiciel est autorisée.\n Vous pouvez supprimer le fichier pid_projet_janus.tmp pour forcer l'execution  \n -- > Ce fichier se trouve dans le meme dossier que l'application <-- ")
            # Quitter l'application :
            print("Une instance du logiciel est déjà en cours d'exécution.")
            logging.info("Une instance du logiciel est déjà en cours d'exécution")
            QtWidgets.QApplication.quit()
            return  # Quitter la fonction si le fichier existe
            

        # Créer le fichier s'il n'existe pas
        with open(filename, "w") as f:
            f.write("PID Projet Janus")
        print("Continuer button clicked")
        logging.info("Continuer button clicked")

        
        # Créez une instance de la classe Ui_Dialog
        dialog = QtWidgets.QDialog(self)
        dialog_ui = Ui_Dialog()
        dialog_ui.setupUi1(dialog)
    
        # On ajoute les boutons du fichier 'gestion_paie_ui_dialog' :
        _translate = QtCore.QCoreApplication.translate
        dialog_ui.pushButton_3.setText(_translate("Dialog", "->Executer <-"))                       # ->Executer <-
        dialog_ui.pushButton_3.clicked.connect(self.Executer)                                       # ->Executer <-
        dialog_ui.pushButton_2.setText(_translate("Dialog", "Quitter"))                             # Quitter
        dialog_ui.pushButton_2.clicked.connect(self.Quitter)                                        # Quitter
        dialog_ui.pushButton.setText(_translate("Dialog", "Gestion des agences"))                   # Gestion des agences
        dialog_ui.pushButton.clicked.connect(self.GestionDesAgences)                                # Gestion des agences
        dialog_ui.pushButton_Cotisation.setText(_translate("Dialog", "..."))                        # Cotisation
        dialog_ui.pushButton_Cotisation.clicked.connect(lambda: self.Cotisation_file(dialog_ui))    # Cotisation
        dialog_ui.pushButton_Matricule.setText(_translate("Dialog", "..."))                         # Matricule
        dialog_ui.pushButton_Matricule.clicked.connect(lambda: self.Matricule_file(dialog_ui))      # Matricule
        dialog_ui.pushButton_Rubrique.setText(_translate("Dialog", "..."))                          # Rubrique
        dialog_ui.pushButton_Rubrique.clicked.connect(lambda: self.Rubrique_file(dialog_ui))        # Rubrique
        dialog_ui.pushButton_Sortie.setText(_translate("Dialog", "..."))                            # Fichier de sortie
        dialog_ui.pushButton_Sortie.clicked.connect(lambda: self.Sortie_file(dialog_ui))            # Fichier de sortie
        # Fichiers charges avec succes: 
        dialog_ui.IF_OK_excel.setVisible(False)
        dialog_ui.IF_OK_excel_2.setVisible(False)
        dialog_ui.IF_OK_excel_3.setVisible(False)
        dialog_ui.IF_OK_excel_4.setVisible(False)
        # Liste des agences:  
        #dialog_ui.comboBox.currentIndexChanged.connect(lambda index, dialog_ui=self.dialog_ui: self.on_index_changed(index))
        dialog_ui.comboBox.currentIndexChanged.connect(lambda index, dialog_ui=dialog_ui: self.on_index_changed(index, dialog_ui))
        # Affichez la boîte de dialogue
        dialog.exec_()

    
    
    def fichier(self):
        ######################
        # Fichier cotisation #
        ######################
        try:
            # Chargement du fichier dans un dataframe 'cotisation' :
            cotisation = pd.read_excel(self.fileName_Cotisation, engine='openpyxl')
        except Exception as e:
            print("Une erreur s'est produite lors de la lecture du fichier Cotisation :", str(e))
            logging.info("Une erreur s'est produite lors de la lecture du fichier Cotisation")
            self.show_error_dialog(f"Une erreur s'est produite lors de la lecture du fichier Cotisation {str(e)} ")
            return     

        ######################
        # Fichier Rubrique   #
        ######################
        try:
            rubrique = pd.read_excel(self.fileName_Rubrique, engine='openpyxl')
        except Exception as e:
            print("Une erreur s'est produite lors de la lecture du fichier Rubrique :", str(e))
            logging.info("Une erreur s'est produite lors de la lecture du fichier Rubrique")
            self.show_error_dialog(f"Une erreur s'est produite lors de la lecture du fichier Rubrique {str(e)} ")
            return rubrique
        
        ######################
        # Fichier Matricule  #
        ######################
        try:
            matricule = pd.read_excel(self.fileName_Matricule, engine='openpyxl')
            # Vérification des colonnes
        except Exception as e:
            print("Une erreur s'est produite lors de la lecture du fichier Matricule :", str(e))
            logging.info("Une erreur s'est produite lors de la lecture du fichier Matricule")
            self.show_error_dialog(f"Une erreur s'est produite lors de la lecture du fichier Matricule {str(e)} ")
            return
        
        #########################
        # Fichier Controle paie #
        #########################
        #
        try:
            # Chargement du fichier dans un dataframe 'controle_paie' en spécifiant la feuille 'Tempo-Banco' : 
            self.controle_paie = pd.read_excel(self.fileName_Sortie, sheet_name='Tempo-Banco', header=None, engine='openpyxl')
            # Assigner le nom du fichier à excel_file_3
            excel_file_3 = self.fileName_Sortie  
            # Obtenir le chemin absolu du fichier
            excel_file_3_path = os.path.abspath(self.fileName_Sortie)
            # Retourner excel_file_3 pour le rendre accessible à l'extérieur de cette méthode
            return excel_file_3  
                
        except Exception as e:
            print("Une erreur s'est produite lors de la lecture du fichier final :", str(e))
            self.show_error_dialog("Une erreur s'est produite lors de la lecture du fichier final :", str(e))
            return None

        ############################    
        # Lancement du programme : #
        ############################
        logging.info("############################ ")
        logging.info("# Lancement du programme : #")
        logging.info("############################")

    def Executer(self, fichier):

        print("Executer pressed")
        logging.info("Executer pressed")

        # Cache le bouton pendant l'execution afin d'empecher 2 lancements simultanes :
        self.dialog_ui.pushButton_3.setEnabled(False)

        # Vérification de la présence des fichiers
        if not (self.fileName_Cotisation and self.fileName_Matricule and self.fileName_Rubrique and self.fileName_Sortie):
            print("Veuillez d'abord sélectionner les fichiers Cotisation, Matricule et Rubrique et Sortie.")
            logging.info("Veuillez d'abord sélectionner les fichiers Cotisation, Matricule et Rubrique et Sortie.")
            self.show_error_dialog("Veuillez d'abord sélectionner les fichiers Cotisation, Matricule et Rubrique et Sortie.")
            return
        #############################
        # Operations sur les data   #
        #############################
        logging.info("#############################")
        logging.info("# Operations sur les data   #")
        logging.info("#############################")

        # Fichier cotisation  # 
        cotisation = pd.read_excel(self.fileName_Cotisation, engine='openpyxl')
        # Selection des lignes de la colonne 'Rubr' qui contiennent uniquement 4 digits :
        rubr = cotisation[cotisation['Rubr'].astype(str).str.match(r'^\d{4}$')]
        # Selection des colonnes 'Mont_Base ', 'Patronal Mont', 'Salarié Mont':
        rubr = rubr[['Rubr', 'Mont_Base ', 'Patronal Mont', 'Salarié Mont']]


        # Fichier Rubrique   #
        rubrique = pd.read_excel(self.fileName_Rubrique, engine='openpyxl')
        rubrique_rubrique = rubrique[rubrique['Rubrique'].astype(str).str.match(r'^\d{4}$')]
        rubrique_rubrique = rubrique_rubrique[['Rubrique', 'Base   ', 'à Payer  ', 'à Retenir  ']]

        # Fichier Matricule  #
        matricule = pd.read_excel(self.fileName_Matricule, engine='openpyxl')
        if self.selected_text == "LILLEGENERALISTE":
            matricule = matricule[matricule['Nom'] == 'Agence'][['Brut', 'Tranche A', 'Tranche B', 'Net Payer', 'Hrs Trav ']]
        else :
            matricule = matricule[matricule['Nom'] == 'Agence'][['Brut', 'Tranche A', 'Tranche B', 'Net Payer', 'Hrs Travaillées']]
        
        # Fichier Controle paie #
        controle_paie = pd.read_excel(self.fileName_Sortie, sheet_name='Tempo-Banco', header=None, engine='openpyxl')
        # Recherche dans l'index de la ligne contenant 'Journal cotisations' : 
        numero_ligne = controle_paie.index[controle_paie.apply(lambda row: 'Journal cotisations' in str(row), axis=1)].tolist()[0]
        # Accéder à la ligne 'Journal cotisations' :
        journal_cotisation = controle_paie.iloc[numero_ligne]
        # Mise en memoire des 60 lignes suivantes pour chaque colonne :
        lignes_suivantes_colonne1 = controle_paie.iloc[numero_ligne+1:numero_ligne+60, 0]  # Première colonne
        lignes_suivantes_colonne2 = controle_paie.iloc[numero_ligne+1:numero_ligne+60, 1]  # Deuxième colonne
        lignes_suivantes_colonne3 = controle_paie.iloc[numero_ligne+1:numero_ligne+60, 2]  # Troisième colonne
        lignes_suivantes_colonne1_base_filtre = controle_paie[controle_paie[0].astype(str).str.contains(r'\b\d{4}\b')]
        # Uniquement les colonnes 0, 1 et 2 nous interessent. 
        lignes_suivantes_colonne1_base_filtre = lignes_suivantes_colonne1_base_filtre[[0, 1, 2]]
        # Utiliser une expression régulière pour extraire les nombres de la colonne 0
        lignes_suivantes_colonne1_base_filtre[0] = lignes_suivantes_colonne1_base_filtre[0].str.extract(r'(\d+)')
        # Renommer la colonne 0 en "Rubr"
        lignes_suivantes_colonne1_base_filtre = lignes_suivantes_colonne1_base_filtre.rename(columns={0:'Rubr'})
        # Renommer la colonne 1 en "Titre"
        lignes_suivantes_colonne1_base_filtre = lignes_suivantes_colonne1_base_filtre.rename(columns={1:'Titre'})
        # Renommer la colonne 2 en "A Remplir"
        lignes_suivantes_colonne1_base_filtre = lignes_suivantes_colonne1_base_filtre.rename(columns={2:'A Remplir'})
        # Convertir la colonne 'Rubr' en chaînes de caractères et nettoyer les espaces
        rubr['Rubr'] = rubr['Rubr'].astype(str).str.strip()
        lignes_suivantes_colonne1_base_filtre['Rubr'] = lignes_suivantes_colonne1_base_filtre['Rubr'].astype(str).str.strip()
        # La variable est correctement définie sur l'attribut de classe
        self.lignes_suivantes_colonne1_base_filtre = lignes_suivantes_colonne1_base_filtre

        # Vérification si lignes_suivantes_colonne1_base_filtre est définie
        if self.lignes_suivantes_colonne1_base_filtre is None:
            print("Veuillez d'abord charger les fichiers.")
            logging.info("Veuillez d'abord charger les fichiers.")
            return

        #################################
        # Fusion de COTISATION et Final #
        #################################
        #
        logging.info("#################################")
        logging.info("# Fusion de COTISATION et Final #")
        logging.info("#################################")
        #
        # Fusionner les DataFrames lignes_suivantes_colonne1_base_filtre et rubr sur la colonne 'Rubr'
        resultat_fusion = pd.merge(self.lignes_suivantes_colonne1_base_filtre, rubr, on='Rubr', how='left')
        # Retirer les lignes avec des valeurs nulles dans la colonne 'Titre'
        resultat_fusion = resultat_fusion.dropna(subset=['Titre'])
        # Reorganisation des collones 
        resultat_fusion = resultat_fusion[['Rubr', 'Titre', 'Mont_Base ', 'Patronal Mont', 'Salarié Mont', 'A Remplir']]
        # Debug
        print("Fusion de COTISATION et Final OK!")
        logging.info("Fusion de COTISATION et Final OK!")

        
        ##############################################
        # Ajout des data de 'Matricule' a la fusion  #
        ##############################################
        #
        logging.info("##############################################")
        logging.info("# Ajout des data de 'Matricule' a la fusion  #")
        logging.info("##############################################")
        #
        # Fusionner resultat_fusion et matricule sur une colonne commune
        resultat_fusion_mat = pd.concat([resultat_fusion, matricule])
        # Reorganisation des collones 
        if self.selected_text == "LILLEGENERALISTE":
            resultat_fusion_mat = resultat_fusion_mat[['Rubr', 'Titre', 'Mont_Base ', 'Patronal Mont', 'Salarié Mont', 'Brut', 'Tranche A', 'Tranche B', 'Net Payer', 'Hrs Trav ', 'A Remplir']]
        else :
            resultat_fusion_mat = resultat_fusion_mat[['Rubr', 'Titre', 'Mont_Base ', 'Patronal Mont', 'Salarié Mont', 'Brut', 'Tranche A', 'Tranche B', 'Net Payer', 'Hrs Travaillées', 'A Remplir']]
        
        # Debug
        print("Ajout des data de 'Matricule' a la fusion OK!")
        print(resultat_fusion_mat)
        logging.info("Ajout des data de 'Matricule' a la fusion OK!")
        logging.info(resultat_fusion_mat)

        ###########################################
        # Ajout des data de 'Rubrique' la fusion  #
        ###########################################
        #
        logging.info("###########################################")
        logging.info("# Ajout des data de 'Rubrique' la fusion  #")
        logging.info("###########################################")
        #
        # Renommer la colonne 'Rubrique' en 'Rubr'
        rubrique_rubrique = rubrique_rubrique.rename(columns={'Rubrique':'Rubr'})
        # Nettoyage des data de la collone 'Rubr'
        rubrique_rubrique['Rubr'] = rubrique_rubrique['Rubr'].astype(str).str.strip()
        # Fusionner les DataFrames resultat_fusion_mat et rubrique_rubrique sur la colonne 'Rubr'
        resultat_fusion_mat_rub = pd.merge(resultat_fusion_mat, rubrique_rubrique, on='Rubr', how='left')

        # Exclure la ligne contenant '1175' dans la colonne 'Rubr' si self.selected_text == "LENS" ou 'BETHUNE"
        if self.selected_text == "LENS":
            resultat_fusion_mat_rub = resultat_fusion_mat_rub[resultat_fusion_mat_rub['Rubr'] != '1175']
        if self.selected_text == "BETHUNE":
            resultat_fusion_mat_rub = resultat_fusion_mat_rub[resultat_fusion_mat_rub['Rubr'] != '1175']

        # Reorganisation des collones 
        if self.selected_text == "LILLEGENERALISTE":
            resultat_fusion_mat_rub= resultat_fusion_mat_rub[['Rubr', 'Titre', 'Mont_Base ', 'Patronal Mont', 'Salarié Mont', 'Brut', 'Tranche A', 'Tranche B', 'Net Payer', 'Hrs Trav ' , 'Base   ', 'à Payer  ', 'à Retenir  ', 'A Remplir']]
        else :
            resultat_fusion_mat_rub= resultat_fusion_mat_rub[['Rubr', 'Titre', 'Mont_Base ', 'Patronal Mont', 'Salarié Mont', 'Brut', 'Tranche A', 'Tranche B', 'Net Payer', 'Hrs Travaillées' , 'Base   ', 'à Payer  ', 'à Retenir  ', 'A Remplir']]
        # Debug
        print("Ajout des data de 'Rubrique' la fusion OK!")
        print(resultat_fusion_mat_rub)
        logging.info("Ajout des data de 'Rubrique' la fusion OK!")
        logging.info(resultat_fusion_mat_rub)
        



        # Cas specifique des lignes qui ne commencent pas par 4 digits 
        logging.info("Cas specifique des lignes qui ne commencent pas par 4 digits ")
        
        #########################################################################################
        #  'Brut total' 'Brut tranche A' 'Brut tranche B' 'Heures travaillées' et 'Net à payer' #
        #########################################################################################
        logging.info("#########################################################################################")
        logging.info("#  'Brut total' 'Brut tranche A' 'Brut tranche B' 'Heures travaillées' et 'Net à payer' #")
        logging.info("#########################################################################################")


        # Recherche dans la page "Tempo-Banco" la ligne contenant 'Journal matricule' : 
        # Ajout du 05/05 cas particulier pour l'agence d'Arras:
        if self.selected_text == "ARRAS":
            ligne_Jmatricule = controle_paie.index[controle_paie.apply(lambda row: 'Journal de matricule' in str(row), axis=1)].tolist()[0]
        # Ajout du 0*/05 cas particulier pour l'agence de Maubeuge:
        #elif self.selected_text == "MAUBEUGE":
        #    ligne_Jmatricule = controle_paie.index[controle_paie.apply(lambda row: 'Journal de matricule' in str(row), axis=1)].tolist()[0]
        else :
            ligne_Jmatricule = controle_paie.index[controle_paie.apply(lambda row: 'Journal matricule' in str(row), axis=1)].tolist()[0]


        # Accéder à la ligne 'Journal cotisations' :
        journal_matricule = controle_paie.iloc[ligne_Jmatricule]
        # Uniquement les colonnes 0, 1 et 2 nous interessent. 
        # Mise en memoire des 6 lignes suivantes pour chaque colonne :
        lignes_suivantes_journal_matricule_colonne1 = controle_paie.iloc[ligne_Jmatricule+1:ligne_Jmatricule+6, 0]  # Première colonne
        lignes_suivantes_journal_matricule_colonne2 = controle_paie.iloc[ligne_Jmatricule+1:ligne_Jmatricule+6, 2]  # Deuxiemme colonne
        lignes_suivantes_journal_matricule_colonne3 = controle_paie.iloc[ligne_Jmatricule+1:ligne_Jmatricule+6, 3]  # Troisième colonne
        
        # Créer un DataFrame temporaire pour chaque colonne
        df_colonne1 = pd.DataFrame(lignes_suivantes_journal_matricule_colonne1)
        df_colonne2 = pd.DataFrame(lignes_suivantes_journal_matricule_colonne2)
        df_colonne3 = pd.DataFrame(lignes_suivantes_journal_matricule_colonne3)

        # Renommer les colonnes pour éviter les conflits lors de la concaténation
        df_colonne1.columns = ['Titre']
        df_colonne2.columns = ['Rubr']
        df_colonne3.columns = ['A Remplir']

        # Concaténer les 3 DataFrames sur l'axe des colonnes
        resultat_fusion_2 = pd.concat([df_colonne1, df_colonne2, df_colonne3], axis=1)
        # Réorganiser les colonnes dans l'ordre original
        resultat_fusion_2= resultat_fusion_2[['Rubr', 'Titre', 'A Remplir']]
        resultat_fusion_2 = pd.concat([resultat_fusion_2], axis=1)
 
        # On va remplir la colonne Rubr avec des chiffres '9900->Brut total', '9901->Brut tranche A'.
        # '9902-> Brut tranche B', '9903->Heures travaillées' et '9904->Net à payer': 
        # Remplacer les valeurs NaN dans la colonne 'Rubr' avec les valeurs spécifiées
        rubr_values = {
            'Rubr': ['9900', '9901', '9902', '9903', '9904'],
            'Titre': ['Brut total', 'Brut tranche A', 'Brut tranche B', 'Heures travaillées', 'Net à payer'],
            'A Remplir': [np.nan, np.nan, np.nan, np.nan, np.nan]
        }
        resultat_fusion_2 = pd.DataFrame(rubr_values)
        resultat_fusion_2= resultat_fusion_2[['Rubr', 'Titre', 'A Remplir']]

        print("resultat_fusion_2")
        print(resultat_fusion_2)

        logging.info("resultat_fusion_2")
        logging.info(resultat_fusion_2)
        
               
        #         Rubr               Titre  A Remplir
        #         9900          Brut total        NaN
        #         9901      Brut tranche A        NaN
        #         9902      Brut tranche B        NaN
        #         9903  Heures travaillées        NaN
        #         9904         Net à payer        NaN

        # On repere les lignes dont on a besoin:
        # la ligne où 'Rubr' est égal à '9000'
        index_9900 = resultat_fusion_2[resultat_fusion_2['Rubr'] == '9900'].index
        # la ligne où 'Rubr' est égal à '9001'
        index_9901 = resultat_fusion_2[resultat_fusion_2['Rubr'] == '9901'].index
        # la ligne où 'Rubr' est égal à '9002'
        index_9902 = resultat_fusion_2[resultat_fusion_2['Rubr'] == '9902'].index
        # la ligne où 'Rubr' est égal à '9003'
        index_9903 = resultat_fusion_2[resultat_fusion_2['Rubr'] == '9903'].index
        # la ligne où 'Rubr' est égal à '9004'
        index_9904 = resultat_fusion_2[resultat_fusion_2['Rubr'] == '9904'].index

        # On copie la valeur de la derniere ligne de 'resultat_fusion_mat'
        # Remplacer la valeur de 'A Remplir' avec la valeur de 'Brut' de la dernière ligne de 'resultat_fusion_mat'
        resultat_fusion_2.loc[index_9900, 'A Remplir'] = resultat_fusion_mat['Brut'].iloc[-1]
        resultat_fusion_2.loc[index_9901, 'A Remplir'] = resultat_fusion_mat['Tranche A'].iloc[-1]
        
        valeur_tranche_b = resultat_fusion_mat['Tranche B'].iloc[-1]
        try:
            if pd.notnull(valeur_tranche_b):
                resultat_fusion_2.loc[index_9902, 'A Remplir'] = float(valeur_tranche_b)
            else:
                resultat_fusion_2.loc[index_9902, 'A Remplir'] = ""
        except (ValueError, TypeError):
            resultat_fusion_2.loc[index_9902, 'A Remplir'] = valeur_tranche_b
            
        if self.selected_text == "LILLEGENERALISTE":
            resultat_fusion_2.loc[index_9903, 'A Remplir'] = resultat_fusion_mat['Hrs Trav '].iloc[-1]
        else: 
            resultat_fusion_2.loc[index_9903, 'A Remplir'] = resultat_fusion_mat['Hrs Travaillées'].iloc[-1]
        resultat_fusion_2.loc[index_9904, 'A Remplir'] = resultat_fusion_mat['Net Payer'].iloc[-1]


        #  Rubr               Titre A Remplir
        #  9900          Brut total   99384.7
        #  9901      Brut tranche A   99384.7
        #  9902      Brut tranche B          
        #  9903  Heures travaillées   5583.76
        #  9904         Net à payer   78018.7
        
        # Ecriture des donnes plus loin


        ############################################################
        # Ligne 'Réduction Chg Tepa Pat' et 'Janus SAS (charges)'  #
        ############################################################

        logging.info("############################################################")
        logging.info("# Ligne 'Réduction Chg Tepa Pat' et 'Janus SAS (charges)'  #")
        logging.info("############################################################")

        # 'Janus SAS (charges)'
        # ======================
        # Recherche dans la page "Tempo-Banco" la ligne contenant 'Agence patronal montant' : 
        ligne_AgencePatronalMontant = controle_paie.index[controle_paie.apply(lambda row: 'Agence patronal montant' in str(row), axis=1)].tolist()[0]
        print("ligne_AgencePatronalMontant: ")
        print(ligne_AgencePatronalMontant)
        # Accéder à la ligne 'Agence patronal montant' :
        agence_patronal_montant = controle_paie.iloc[ligne_AgencePatronalMontant][:3]

        # Créer un DataFrame à partir de la ligne "Agence patronal montant"
        resultat_fusion_4 = pd.DataFrame({
            'Rubr': [agence_patronal_montant[0]],
            'Titre': [agence_patronal_montant[1]],
            'A Remplir': [agence_patronal_montant[2]]
        })

        print("resultat_fusion_4: ")
        print(resultat_fusion_4)

        logging.info("resultat_fusion_4: ")
        logging.info(resultat_fusion_4)

        # Importation des valeurs à remplir depuis le fichier Cotisation
        # Nous recherchons la lignes dont la colonne 'Rubrique' comporte la valeur 'Agence'
        cotisation_agence = pd.read_excel(self.fileName_Cotisation, engine='openpyxl')
        # Filtrer la lignes où la colonne 'Rubrique' contient la valeur 'Agence'
        cotisation_agence = cotisation_agence[cotisation_agence['Rubr'].str.contains('Agence', case=False, na=False)]
        print("cotisation_agence: ")
        print(cotisation_agence)

        logging.info("cotisation_agence: ")
        logging.info(cotisation_agence)

        # Obtenir la valeur de la colonne 'Patronal Mont' de la première ligne de 'cotisation_agence'
        patronal_mont_value = cotisation_agence.iloc[0]['Patronal Mont']

        # Affecter cette valeur à la colonne 'A Remplir' de 'resultat_fusion_4'
        resultat_fusion_4['A Remplir'] = patronal_mont_value

        print("resultat_fusion_4 avec la valeur de 'Patronal Mont' de 'cotisation_agence' : ")
        print(resultat_fusion_4)

        logging.info("resultat_fusion_4 avec la valeur de 'Patronal Mont' de 'cotisation_agence' : ")
        logging.info(resultat_fusion_4)



        ####################################################
        #    lignes 'BRUT à payer', 'Fiscal' et 'Net ***'  #
        ####################################################
        logging.info("####################################################")
        logging.info("#    lignes 'BRUT à payer', 'Fiscal' et 'Net ***'  #")
        logging.info("####################################################")

        # Recherche dans la page "Tempo-Banco" la ligne contenant 'Journal de rubriques' : 
        ligne_Jrubriques = controle_paie.index[controle_paie.apply(lambda row: 'Journal de rubriques' in str(row), axis=1)].tolist()[0]
        # Accéder à la ligne 'Journal de rubriques' :
        journal_rubriques = controle_paie.iloc[ligne_Jrubriques]
        # Uniquement les colonnes 0, 1 et 2 nous interessent. 
        # Mise en memoire des 25 lignes suivantes pour chaque colonne :
        lignes_suivantes_journal_rubriques_colonne1 = controle_paie.iloc[ligne_Jrubriques+1:ligne_Jrubriques+25, 0]  # Première colonne
        lignes_suivantes_journal_rubriques_colonne2 = controle_paie.iloc[ligne_Jrubriques+1:ligne_Jrubriques+25, 1]  # Deuxiemme colonne
        lignes_suivantes_journal_rubriques_colonne3 = controle_paie.iloc[ligne_Jrubriques+1:ligne_Jrubriques+25, 2]  # Troisième colonne

        # Créer un DataFrame temporaire pour chaque colonne
        df_colonne1_1 = pd.DataFrame(lignes_suivantes_journal_rubriques_colonne1)
        df_colonne2_1 = pd.DataFrame(lignes_suivantes_journal_rubriques_colonne2)
        df_colonne3_1 = pd.DataFrame(lignes_suivantes_journal_rubriques_colonne3)       

        # Renommer les colonnes pour éviter les conflits lors de la concaténation
        df_colonne2_1.columns = ['Titre']
        df_colonne1_1.columns = ['Rubr']
        df_colonne3_1.columns = ['A Remplir']

        # Concaténer les 3 DataFrames sur l'axe des colonnes
        resultat_fusion_3 = pd.concat([df_colonne1_1, df_colonne2_1, df_colonne3_1], axis=1)
        # Réorganiser les colonnes dans l'ordre original
        resultat_fusion_3= resultat_fusion_3[['Rubr', 'Titre', 'A Remplir']]
        resultat_fusion_3 = pd.concat([resultat_fusion_3], axis=1)

        # Debug: Afficher TOUT le contenu de resultat_fusion_3 pour voir ce qui existe
        print("=== DEBUG COMPLET resultat_fusion_3 INITIAL ===")
        print("Toutes les lignes de resultat_fusion_3:")
        for idx, row in resultat_fusion_3.iterrows():
            print(f"Index {idx}: Rubr='{row['Rubr']}', Titre='{row['Titre']}', A Remplir='{row['A Remplir']}'")
        print("=== FIN DEBUG COMPLET ===")
        logging.info("DEBUG COMPLET resultat_fusion_3:")
        logging.info(resultat_fusion_3)


        # Recherche dans le DataFrame fusionné la ligne dont le Titre contient 'BRUT à payer'
        ligne_BRUT_a_payer = resultat_fusion_3[resultat_fusion_3['Titre'].str.contains('BRUT à payer', case=False, na=False)]
        # Recherche dans le DataFrame fusionné la ligne dont le Titre contient 'Fiscal'
        ligne_Fiscal = resultat_fusion_3[resultat_fusion_3['Titre'].str.contains('Fiscal', case=False, na=False)]
        # Recherche dans le DataFrame fusionné la ligne dont le Titre contient 'Net ***' pour Total à payer
        ligne_Total_a_payer = resultat_fusion_3[resultat_fusion_3['Titre'].str.contains('Net.*\*', case=False, na=False, regex=True)]
        
        # Si aucune ligne "Net ***" n'est trouvée, créer une ligne pour "Total à payer"
        if ligne_Total_a_payer.empty:
            print("Aucune ligne 'Net ***' trouvée dans resultat_fusion_3. Création manuelle d'une ligne 'Total à payer'.")
            logging.info("Aucune ligne 'Net ***' trouvée dans resultat_fusion_3. Création manuelle d'une ligne 'Total à payer'.")
            
            # Créer une ligne vide pour "Total à payer"
            nouvelle_ligne = pd.DataFrame({
                'Rubr': ['Total à payer'],
                'Titre': ['Net ***'], 
                'A Remplir': [None]
            })
            
            # Ajouter cette ligne à resultat_fusion_3
            resultat_fusion_3 = pd.concat([resultat_fusion_3, nouvelle_ligne], ignore_index=True)
            
            # Récupérer à nouveau la ligne
            ligne_Total_a_payer = resultat_fusion_3[resultat_fusion_3['Titre'].str.contains('Net.*\*', case=False, na=False, regex=True)]
            
            print("Ligne 'Total à payer' créée manuellement:")
            print(ligne_Total_a_payer)
            logging.info("Ligne 'Total à payer' créée manuellement:")
            logging.info(ligne_Total_a_payer)

        # Pour l'agence de Lens la valeur 1175 vaut 1440 donc on va la traiter a part ici, la meme chose pour l'agence de Bethune : 
        if self.selected_text == "LENS":
            # Faire quelque chose si self.selected_text est égal à "LENS"
            print("Pour l'agence de Lens la valeur 1175 vaut 1440 donc on va la traiter a part ici:")
            # Recherche dans le DataFrame fusionné la ligne dont la rubr contient '1175 base'
            ligne_1175 = resultat_fusion_3[resultat_fusion_3['Rubr'].str.contains('1175 base', case=False, na=False)]
            print(" Recherche dans le DataFrame fusionné la ligne dont la rubr contient 1175 base")
            print(ligne_1175)
            # Importation des valeurs à remplir depuis le fichier Rubrique
            # Nous recherchons les lignes dont la colonne 'Rubrique' comporte la valeur '1440'
            rubrique_total_LENS = pd.read_excel(self.fileName_Rubrique, engine='openpyxl')
            rubrique_total_LENS_1440 = rubrique_total_LENS[rubrique_total_LENS['Intitulé'].str.contains('Heures payées non travaillées', case=False, na=False)]
            print("Rubrique Total LENS 1440 :")
            print(rubrique_total_LENS_1440)
            # Copie de la valeur de la colonne 'Base' de 'rubrique_total_LENS_1440' vers la colonne 'A Remplir' de 'ligne_1175' .
            # Assurez-vous qu'une seule ligne est sélectionnée dans rubrique_total_LENS_1440
            if len(rubrique_total_LENS_1440) == 1:
                # Récupérer la valeur de la colonne 'Base   ' de rubrique_total_LENS_1440
                rubrique_total_LENS_1440_Base = rubrique_total_LENS_1440.iloc[0]['Base   ']
                # Modifier la valeur dans la colonne 'A Remplir' de ligne_1175
                ligne_1175.loc[:, 'A Remplir'] = rubrique_total_LENS_1440_Base
                print("Copie de la valeur de la colonne Base de rubrique_total_LENS_1440 vers la colonne A Remplir de ligne_1175")
                print(rubrique_total_LENS_1440_Base)
                logging.info("Copie réussie pour BRUT à payer.")
                # Modifier la valeur dans la colonne 'A Remplir' de ligne_1175
                ligne_1175.loc[:, 'A Remplir'] = rubrique_total_LENS_1440_Base
                print("Copie réussie pour ligne_1175.")
                print(ligne_1175)
                logging.info("Copie réussie pour ligne_1175.")
            else:
                print("Erreur: Plus d'une ligne trouvée dans rubrique_total_Brut. Veuillez vérifier vos données.")
                logging.info("Erreur: Plus d'une ligne trouvée dans rubrique_total_Brut. Veuillez vérifier vos données.")

        if self.selected_text == "BETHUNE":
            # Faire quelque chose si self.selected_text est égal à "BETHUNE"
            print("Pour l'agence de Bethune la valeur 1175 vaut 1425 donc on va la traiter a part ici:")
            # Recherche dans le DataFrame fusionné la ligne dont la rubr contient '1175 base'
            ligne_1175 = resultat_fusion_3[resultat_fusion_3['Rubr'].str.contains('1175 base', case=False, na=False)]
            print(" Recherche dans le DataFrame fusionné la ligne dont la rubr contient 1175 base")
            print(ligne_1175)
            # Importation des valeurs à remplir depuis le fichier Rubrique
            # Nous recherchons les lignes dont la colonne 'Rubrique' comporte la valeur '1425'
            rubrique_total_BETHUNE = pd.read_excel(self.fileName_Rubrique, engine='openpyxl')
            rubrique_total_BETHUNE_1425 = rubrique_total_BETHUNE[rubrique_total_BETHUNE['Intitulé'].str.contains('Heures Habillage', case=False, na=False)]
            print("Rubrique Total BETHUNE 1440 :")
            print(rubrique_total_BETHUNE_1425)
            # Copie de la valeur de la colonne 'Base' de 'rubrique_total_LENS_1440' vers la colonne 'A Remplir' de 'ligne_1175' .
            # Assurez-vous qu'une seule ligne est sélectionnée dans rubrique_total_LENS_1440
            if len(rubrique_total_BETHUNE_1425) == 1:
                # Récupérer la valeur de la colonne 'Base   ' de rubrique_total_LENS_1440
                rubrique_total_BETHUNE_1425_Base = rubrique_total_BETHUNE_1425.iloc[0]['Base   ']
                # Modifier la valeur dans la colonne 'A Remplir' de ligne_1175
                ligne_1175.loc[:, 'A Remplir'] = rubrique_total_BETHUNE_1425_Base
                print("Copie de la valeur de la colonne Base de rubrique_total_BETHUNE_1440 vers la colonne A Remplir de ligne_1175")
                print(rubrique_total_BETHUNE_1425_Base)
                logging.info("Copie réussie pour BRUT à payer.")
                # Modifier la valeur dans la colonne 'A Remplir' de ligne_1175
                ligne_1175.loc[:, 'A Remplir'] = rubrique_total_BETHUNE_1425_Base
                print("Copie réussie pour ligne_1175.")
                print(ligne_1175)
                logging.info("Copie réussie pour ligne_1175.")
                logging.info(ligne_1175)
                logging.info("Copie réussie pour rubrique_total_BETHUNE_1425_Base.")
                logging.info(rubrique_total_BETHUNE_1425_Base)

            else:
                print("Erreur: Plus d'une ligne trouvée dans rubrique_total_Brut. Veuillez vérifier vos données.")
                logging.info("Erreur: Plus d'une ligne trouvée dans rubrique_total_Brut. Veuillez vérifier vos données.")


        # Ligne BRUT à payer :
        # Rubr              Titre  A Remplir
        # Total      BRUT à payer        NaN

        # ligne_Fiscal :
        # Rubr              Titre  A Remplir
        # Total            Fiscal        NaN

        # ligne_Net :
        # Rubr              Titre   A Remplir
        # Total à payer  Net ***         NaN

        # Importation des valeurs à remplir depuis le fichier Rubrique
        # Nous recherchons les lignes dont la colonne 'Rubrique' comporte la valeur 'Total'
        rubrique_total = pd.read_excel(self.fileName_Rubrique, engine='openpyxl')

        # Filtrer les lignes où la colonne 'Rubrique' contient la valeur 'Total'
        rubrique_total = rubrique_total[rubrique_total['Rubrique'].str.contains('Total', case=False, na=False)]

        # Debug: Afficher TOUT le contenu de rubrique_total pour voir toutes les lignes disponibles
        print("=== DEBUG COMPLET rubrique_total ===")
        print("Toutes les lignes avec 'Total' dans rubrique_total:")
        for idx, row in rubrique_total.iterrows():
            print(f"Index {idx}: Rubrique='{row['Rubrique']}', Intitulé='{row['Intitulé']}', à Payer={row['à Payer  ']}")
        print("=== FIN DEBUG COMPLET rubrique_total ===")

        # Afficher les valeurs importées
        print("Rubrique Total :")
        print(rubrique_total)

        logging.info("Afficher les valeurs importées")
        logging.info("Rubrique Total :")
        logging.info(rubrique_total)

        # Filtrer la lignes où la colonne 'Intitulé' contient la valeur 'Brut'
        rubrique_total_Brut = rubrique_total[rubrique_total['Intitulé'].str.contains('Brut', case=False, na=False)]

        # Filtrer la lignes où la colonne 'Intitulé' contient la valeur 'Fiscal'
        rubrique_total_Fiscal = rubrique_total[rubrique_total['Intitulé'].str.contains('Fiscal', case=False, na=False)]

        # Rechercher la ligne avec Net *** dans l'intitulé (pour Total à payer)
        # Chercher d'abord "Net ***" sans chiffres, sinon prendre n'importe quelle ligne avec "Net"
        rubrique_total_Net = rubrique_total[rubrique_total['Intitulé'].str.contains('Net.*\*', case=False, na=False, regex=True)]
        if rubrique_total_Net.empty:
            # Fallback: chercher toute ligne avec "Net" 
            rubrique_total_Net = rubrique_total[rubrique_total['Intitulé'].str.contains('Net', case=False, na=False)]
        
        # Priorité : prendre la ligne "Net ***" sans chiffres entre parenthèses
        if len(rubrique_total_Net) > 1:
            # Chercher la ligne exacte "Net ***" (sans chiffres entre parenthèses)
            rubrique_net_simple = rubrique_total_Net[~rubrique_total_Net['Intitulé'].str.contains(r'\(\d+\)', case=False, na=False, regex=True)]
            if not rubrique_net_simple.empty:
                rubrique_total_Net = rubrique_net_simple
                print("Sélection de la ligne 'Net ***' sans chiffres entre parenthèses")
            else:
                print("Aucune ligne 'Net ***' simple trouvée, utilisation de la première ligne disponible")

        # Afficher les valeurs importées
        print("rubrique_total_Brut :")
        print(rubrique_total_Brut)
        print("rubrique_total_Fiscal :")
        print(rubrique_total_Fiscal)
        print("rubrique_total_Net :")
        print(rubrique_total_Net)
        print("=== DEBUG TOTAL À PAYER ===")
        print(f"Nombre de lignes trouvées dans rubrique_total_Net: {len(rubrique_total_Net)}")
        if len(rubrique_total_Net) > 0:
            print("Valeurs à payer disponibles:")
            for idx, row in rubrique_total_Net.iterrows():
                print(f"  - Intitulé: '{row['Intitulé']}', à Payer: {row['à Payer  ']}")
        print(f"Nombre de lignes trouvées pour ligne_Total_a_payer: {len(ligne_Total_a_payer)}")
        print("=== FIN DEBUG ===")

        
        logging.info("Afficher les valeurs importées")
        logging.info("rubrique_total_Brut :")
        logging.info(rubrique_total_Brut)
        logging.info("rubrique_total_Fiscal :")
        logging.info(rubrique_total_Fiscal)
        logging.info("rubrique_total_Net :")
        logging.info(rubrique_total_Net)
        #rubrique_total_Brut :
        # Agence Rubrique     Intitulé CptDeb CptCré Nbr Base         à Payer   à Retenir  
        # 22      1    Total  Brut    ***    NaN    NaN               99384.7            
        # rubrique_total_Fiscal :
        # Agence Rubrique    Intitulé CptDeb CptCré Nbr Base          à Payer   à Retenir  
        # 39      1    Total  Fiscal ***    NaN    NaN               78464.64            
        # rubrique_total_Net :
        # Agence Rubrique      Intitulé CptDeb CptCré Nbr Base         à Payer   à Retenir  
        # 51      1    Total  Net (1)  ***    NaN    NaN              85044.57            
        # 56      1    Total    Net    ***    NaN    NaN               78018.7
        
        # La ligne qui contient un chiffe avec Net ne nous interesse pas : 
        # Filtre pour sélectionner la ligne qui ne contient pas de chiffre dans la colonne 'Intitulé'
        # rubrique_total_Net_filtre = rubrique_total_Net[~rubrique_total_Net['Intitulé'].str.contains(r'\d', regex=True, na=False)]
        
        # Copie de la valeur de la colonne 'à Payer  ' de 'rubrique_total_Brut' vers la colonne 'A Remplir' de 'ligne_BRUT_a_payer', 'ligne_Fiscal' et 'ligne_Net'.

        # Assurez-vous qu'une seule ligne est sélectionnée dans rubrique_total_Brut
        if len(rubrique_total_Brut) == 1:
            # Récupérer la valeur de la colonne 'à Payer' de rubrique_total_Brut
            valeur_a_payer = rubrique_total_Brut.iloc[0]['à Payer  ']

            # Modifier la valeur dans la colonne 'A Remplir' de ligne_BRUT_a_payer
            ligne_BRUT_a_payer.loc[:, 'A Remplir'] = valeur_a_payer
            print("Copie réussie pour BRUT à payer.")
            logging.info("Copie réussie pour BRUT à payer.")
        else:
            print("Erreur: Plus d'une ligne trouvée dans rubrique_total_Brut. Veuillez vérifier vos données.")

        # Assurez-vous qu'une seule ligne est sélectionnée dans rubrique_total_Fiscal
        if len(rubrique_total_Fiscal) == 1:
            # Récupérer la valeur de la colonne 'à Payer' de rubrique_total_Fiscal
            valeur_a_payer_fiscal = rubrique_total_Fiscal.iloc[0]['à Payer  ']

            # Modifier la valeur dans la colonne 'A Remplir' de ligne_Fiscal
            ligne_Fiscal.loc[:, 'A Remplir'] = valeur_a_payer_fiscal
            print("Copie réussie pour Fiscal.")
            logging.info("Copie réussie pour Fiscal")
        else:
            print("Erreur: Plus d'une ligne trouvée dans rubrique_total_Fiscal. Veuillez vérifier vos données.")

       
        # Traitement pour "Total à payer" - Utiliser la première ligne disponible de rubrique_total_Net
        if len(rubrique_total_Net) >= 1:
            # Récupérer la valeur de la colonne 'à Payer' de la première ligne disponible
            valeur_a_payer_net = rubrique_total_Net.iloc[0]['à Payer  ']
            # Modifier la valeur dans la colonne 'A Remplir' de ligne_Total_a_payer
            if not ligne_Total_a_payer.empty:
                ligne_Total_a_payer.loc[:, 'A Remplir'] = valeur_a_payer_net
                print(f"Copie réussie pour Total à payer: {valeur_a_payer_net}")
                logging.info(f"Copie réussie pour Total à payer: {valeur_a_payer_net}")
            else:
                print("Erreur: ligne_Total_a_payer est vide")
                logging.error("Erreur: ligne_Total_a_payer est vide")
        else:
            print("Erreur: Aucune ligne 'Net' trouvée dans rubrique_total.")
            logging.error("Erreur: Aucune ligne 'Net' trouvée dans rubrique_total.")

        # Intégration des données de 'ligne_BRUT_a_payer' dans 'resultat_fusion_3'
        if not ligne_BRUT_a_payer.empty:
            brut_a_payer_value = ligne_BRUT_a_payer.iloc[0]['A Remplir']
            resultat_fusion_3.loc[resultat_fusion_3['Titre'].str.contains('BRUT à payer', case=False, na=False), 'A Remplir'] = brut_a_payer_value

        # Intégration des données de 'ligne_Fiscal' dans 'resultat_fusion_3'
        if not ligne_Fiscal.empty:
            fiscal_value = ligne_Fiscal.iloc[0]['A Remplir']
            resultat_fusion_3.loc[resultat_fusion_3['Titre'].str.contains('Fiscal', case=False, na=False), 'A Remplir'] = fiscal_value

        # Intégration des données de 'ligne_Total_a_payer' dans 'resultat_fusion_3'
        if not ligne_Total_a_payer.empty:
            total_a_payer_value = ligne_Total_a_payer.iloc[0]['A Remplir']
            # Mettre à jour toutes les lignes contenant "Net ***" ou "Total à payer"
            mask_net = resultat_fusion_3['Titre'].str.contains('Net.*\*', case=False, na=False, regex=True)
            mask_total = resultat_fusion_3['Titre'].str.contains('Total à payer', case=False, na=False)
            resultat_fusion_3.loc[mask_net | mask_total, 'A Remplir'] = total_a_payer_value
            print(f"Intégration Total à payer réussie avec la valeur: {total_a_payer_value}")
            logging.info(f"Intégration Total à payer réussie avec la valeur: {total_a_payer_value}")
        else:
            print("Erreur: ligne_Total_a_payer est vide, impossible d'intégrer la valeur")
            logging.error("Erreur: ligne_Total_a_payer est vide, impossible d'intégrer la valeur")

        if self.selected_text == "LENS":
            if not ligne_1175.empty:
                l1175_value = ligne_1175.iloc[0]['A Remplir']
                resultat_fusion_3.loc[resultat_fusion_3['Titre'].str.contains('VM', case=False, na=False), 'A Remplir'] = l1175_value

        if self.selected_text == "BETHUNE":
            if not ligne_1175.empty:
                l1175_value = ligne_1175.iloc[0]['A Remplir']
                resultat_fusion_3.loc[resultat_fusion_3['Titre'].str.contains('Heures visites médicales', case=False, na=False), 'A Remplir'] = l1175_value


        # Afficher les données de 'resultat_fusion_3' après intégration
        # print("resultat_fusion_3 après intégration :")
        # print(resultat_fusion_3)

        # Maintenant, on conserve uniquement les lignes qui contiennent une valeur non null dans la colonne 'A Remplir' de 'resultat_fusion_3':
        # Supprimer les lignes avec des valeurs nulles dans la colonne 'A Remplir'
        # resultat_fusion_3 = resultat_fusion_3.dropna(subset=['A Remplir'])

        # Afficher les données de 'resultat_fusion_3' après la suppression des lignes avec des valeurs nulles
        # print("resultat_fusion_3 après suppression des lignes avec des valeurs nulles :")
        print("=== CONTENU FINAL DE resultat_fusion_3 ===")
        print(resultat_fusion_3)
        print("Lignes contenant 'Net' ou 'Total à payer':")
        mask_debug = resultat_fusion_3['Titre'].str.contains('Net|Total.*payer', case=False, na=False, regex=True)
        print(resultat_fusion_3[mask_debug])
        print("=== FIN CONTENU FINAL ===")
        logging.info("resultat_fusion_3")
        logging.info(resultat_fusion_3)

        # resultat_fusion_3 après suppression des lignes avec des valeurs nulles :
        #              Rubr         Titre  A Remplir
        # 78         Total   BRUT à payer   99384.70
        # 84          Total        Fiscal   78464.64
        # 92  Total à payer       Net ***   78018.70
        
        ############################################################################
        # Controle de l'agence afin d'avoir les equivalents 'Rubr' ==> 'A Remplir' #
        ############################################################################
        #
        logging.info("############################################################################")
        logging.info(" # Controle de l'agence afin d'avoir les equivalents 'Rubr' ==> 'A Remplir'#")
        logging.info("############################################################################")

        # Ici on associe va deplacer  les donnees de 'A Remplir' en focntion des codes utilises dans chaque rubrique      
        
        # Remplace les données en fonction des correspondances de gestion_des_agences:
                
        for rubrique, colonne in self.correspondances_selectionnees.items():
            resultat_fusion_mat_rub.loc[resultat_fusion_mat_rub['Rubr'] == rubrique, 'A Remplir'] = resultat_fusion_mat_rub.loc[resultat_fusion_mat_rub['Rubr'] == rubrique, colonne]
        
               
        

        print("Remplace les données en fonction des correspondances de gestion_des_agences:")
        print(resultat_fusion_mat_rub)

        logging.info("Remplace les données en fonction des correspondances de gestion_des_agences:")
        logging.info(resultat_fusion_mat_rub)


        ###########################################
        #   Cas specifique des lignes en doublon  # 
        ###########################################
        logging.info("###########################################")
        logging.info("#   Cas specifique des lignes en doublon  #")
        logging.info("###########################################")
        
        # Filtrer les lignes de resultat_fusion_mat_rub qui sont des doublons
        resultat_fusion_mat_rub_doublons = resultat_fusion_mat_rub[resultat_fusion_mat_rub.duplicated(subset=['Rubr'], keep=False)]

        # Ajouute un peu de debug:
        print("Filtrer les lignes de resultat_fusion_mat_rub qui sont des doublons :")
        print(resultat_fusion_mat_rub_doublons)


        logging.info("Filtrer les lignes de resultat_fusion_mat_rub qui sont des doublons :")
        logging.info(resultat_fusion_mat_rub_doublons)


       
        resultat_fusion_mat_rub_doublons_null = resultat_fusion_mat_rub_doublons

        
        # Compter le nombre d'occurrences de chaque valeur dans la colonne 'Rubr'
        rubrique_counts = resultat_fusion_mat_rub_doublons_null['Rubr'].value_counts()

        print("Filtrer les lignes qui apparaissent 2 fois avec la même valeur dans la colonne Rubr :")
        # Filtrer les lignes qui apparaissent 2 fois avec la même valeur dans la colonne 'Rubr'
        rubriques_doubles = rubrique_counts[rubrique_counts == 2].index
        print("Filtrer les lignes qui apparaissent 2 fois avec la même valeur dans la colonne Rubr: ")
        print(rubriques_doubles)
        
        logging.info("Filtrer les lignes qui apparaissent 2 fois avec la même valeur dans la colonne Rubr: ")
        logging.info(rubriques_doubles)

        # Vérifier si self.rubriques_en_doublon_selectionnees est défini
        if self.rubriques_en_doublon_selectionnees:
            # Récupérer les valeurs du dictionnaire self.rubriques_en_doublon_selectionnees
            valeurs = list(self.rubriques_en_doublon_selectionnees.values())

            nombre_de_dictionnaires = len(self.rubriques_en_doublon_selectionnees)
            print("Le nombre de dictionnaires contenus dans self.rubriques_en_doublon_test_selectionnees est :", nombre_de_dictionnaires)
            logging.info("Le nombre de dictionnaires contenus dans self.rubriques_en_doublon_test_selectionnees est :", nombre_de_dictionnaires)
            # S'assurer qu'il y a des valeurs dans le 1er dictionnaire
            if valeurs[0]:
                if nombre_de_dictionnaires >= 1:
                    # Remplacement de la valeur de Rubr
                    ###################################
                    # Accéder à la première  sous-liste de valeurs
                    sous_liste_valeurs = valeurs[0]  
                    # Extraire les valeurs individuelles dans des variables distinctes
                    valeur_1, valeur_2, valeur_3 = sous_liste_valeurs
                    # Afficher les valeurs individuelles
                    print("Valeur 1 de la 1er sous liste:", valeur_1)
                    print("Valeur 2 de la 1er sous liste:", valeur_2)
                    print("Valeur 3 de la 1er sous liste:", valeur_3)
                    
                    rubrique_valeur_1 = resultat_fusion_mat_rub_doublons_null.loc[resultat_fusion_mat_rub_doublons_null['Rubr'] == valeur_1]
                    print("Lignes contenant la valeur 1 de la 1er sous liste dans la colonne 'Rubr':")
                    print(rubrique_valeur_1)                
                    rubrique_valeur_1.loc[rubrique_valeur_1.index[0], 'Rubr'] = valeur_2
                    rubrique_valeur_1.loc[rubrique_valeur_1.index[1], 'Rubr'] = valeur_3
                    print("Remplacement de la valeur de Rubr: ")
                    print(rubrique_valeur_1)

                    # Remplacement de la valeur de 'A Remplir'
                    ###########################################
                    if self.rubriques_en_doublon_2_selectionnees:
                        # Récupérer les valeurs du dictionnaire self.rubriques_en_doublon_2_selectionnees
                        valeurs_de_a_remplir = list(self.rubriques_en_doublon_2_selectionnees.values())
                        if valeurs_de_a_remplir[0]:
                            sous_liste_valeurs_0 = valeurs_de_a_remplir[0]
                            sous_liste_valeurs_1 = valeurs_de_a_remplir[1]
                            print("sous_liste_valeurs_0")
                            print(sous_liste_valeurs_0)
                            print("sous_liste_valeurs_1")
                            print(sous_liste_valeurs_1)
                            rubrique_valeur_1.loc[rubrique_valeur_1.index[0], 'A Remplir'] = rubrique_valeur_1.loc[rubrique_valeur_1.index[0], sous_liste_valeurs_0]
                            rubrique_valeur_1.loc[rubrique_valeur_1.index[1], 'A Remplir'] = rubrique_valeur_1.loc[rubrique_valeur_1.index[1], sous_liste_valeurs_1]
                            print("Remplacement de la valeur de A remplir dans le dictionnaire 1 : ")
                            print(rubrique_valeur_1)
                            logging.info("Remplacement de la valeur de A remplir dans le dictionnaire 1 : ")
                            logging.info(rubrique_valeur_1)

                ########################################################################
                # Le meme modele va etre applique 10 fois afin de supporter 10 doublons#
                ########################################################################

                # S'assurer qu'il y a des valeurs dans le 2nd dictionnaire     
                if nombre_de_dictionnaires >= 2:
                    if valeurs[1]:                    
                        sous_liste_valeurs_1 = valeurs[1]
                        valeur_101, valeur_201, valeur_301 = sous_liste_valeurs_1
                        rubrique_valeur_2 = resultat_fusion_mat_rub_doublons_null.loc[resultat_fusion_mat_rub_doublons_null['Rubr'] == valeur_101]
                        rubrique_valeur_2.loc[rubrique_valeur_2.index[0], 'Rubr'] = valeur_201
                        rubrique_valeur_2.loc[rubrique_valeur_2.index[1], 'Rubr'] = valeur_301
                        valeurs_de_a_remplir = list(self.rubriques_en_doublon_2_selectionnees.values())
                        sous_liste_valeurs_2 = valeurs_de_a_remplir[2]
                        sous_liste_valeurs_3 = valeurs_de_a_remplir[3]
                        rubrique_valeur_2.loc[rubrique_valeur_2.index[0], 'A Remplir'] = rubrique_valeur_2.loc[rubrique_valeur_2.index[0], sous_liste_valeurs_2]
                        rubrique_valeur_2.loc[rubrique_valeur_2.index[1], 'A Remplir'] = rubrique_valeur_2.loc[rubrique_valeur_2.index[0], sous_liste_valeurs_3]
                        print("Remplacement de la valeur de A remplir dans le dictionnaire 2 : ")
                        print(rubrique_valeur_2)
                        logging.info("Remplacement de la valeur de A remplir dans le dictionnaire 2 : ")
                        logging.info(rubrique_valeur_2)

                # S'assurer qu'il y a des valeurs dans le 3rd dictionnaire 
                if nombre_de_dictionnaires >= 3:    
                    if valeurs[2]:                    
                        sous_liste_valeurs_102 = valeurs[2]
                        valeur_102, valeur_202, valeur_302 = sous_liste_valeurs_102
                        rubrique_valeur_3 = resultat_fusion_mat_rub_doublons_null.loc[resultat_fusion_mat_rub_doublons_null['Rubr'] == valeur_102]
                        rubrique_valeur_3.loc[rubrique_valeur_3.index[0], 'Rubr'] = valeur_202
                        rubrique_valeur_3.loc[rubrique_valeur_3.index[1], 'Rubr'] = valeur_302
                        valeurs_de_a_remplir = list(self.rubriques_en_doublon_2_selectionnees.values())
                        sous_liste_valeurs_4 = valeurs_de_a_remplir[4]
                        sous_liste_valeurs_5 = valeurs_de_a_remplir[5]
                        rubrique_valeur_3.loc[rubrique_valeur_3.index[0], 'A Remplir'] = rubrique_valeur_3.loc[rubrique_valeur_3.index[0], sous_liste_valeurs_4]
                        rubrique_valeur_3.loc[rubrique_valeur_3.index[1], 'A Remplir'] = rubrique_valeur_3.loc[rubrique_valeur_3.index[0], sous_liste_valeurs_5]
                        print("Remplacement de la valeur de A remplir dans le dictionnaire 3 : ")
                        print(rubrique_valeur_3)
                        logging.info("Remplacement de la valeur de A remplir dans le dictionnaire 3 : ")
                        logging.info(rubrique_valeur_3)

                # S'assurer qu'il y a des valeurs dans le dictionnaire 4   
                if nombre_de_dictionnaires >= 4:  
                    if valeurs[3]:                    
                        sous_liste_valeurs_103 = valeurs[3]
                        valeur_103, valeur_203, valeur_303 = sous_liste_valeurs_103
                        rubrique_valeur_4 = resultat_fusion_mat_rub_doublons_null.loc[resultat_fusion_mat_rub_doublons_null['Rubr'] == valeur_103]
                        rubrique_valeur_4.loc[rubrique_valeur_4.index[0], 'Rubr'] = valeur_203
                        rubrique_valeur_4.loc[rubrique_valeur_4.index[1], 'Rubr'] = valeur_303
                        valeurs_de_a_remplir = list(self.rubriques_en_doublon_2_selectionnees.values())
                        sous_liste_valeurs_6 = valeurs_de_a_remplir[6]
                        sous_liste_valeurs_7 = valeurs_de_a_remplir[7]
                        rubrique_valeur_4.loc[rubrique_valeur_4.index[0], 'A Remplir'] = rubrique_valeur_4.loc[rubrique_valeur_4.index[0], sous_liste_valeurs_6]
                        rubrique_valeur_4.loc[rubrique_valeur_4.index[1], 'A Remplir'] = rubrique_valeur_4.loc[rubrique_valeur_4.index[0], sous_liste_valeurs_7]
                        print("Remplacement de la valeur de A remplir dans le dictionnaire 4 : ")
                        print(rubrique_valeur_4)
                        logging.info("Remplacement de la valeur de A remplir dans le dictionnaire 4 : ")
                        logging.info(rubrique_valeur_4)

                # S'assurer qu'il y a des valeurs dans le dictionnaire 5   
                if nombre_de_dictionnaires >= 5:
                    if valeurs[4]:                    
                        sous_liste_valeurs_104 = valeurs[4]
                        valeur_104, valeur_204, valeur_304 = sous_liste_valeurs_104
                        rubrique_valeur_5 = resultat_fusion_mat_rub_doublons_null.loc[resultat_fusion_mat_rub_doublons_null['Rubr'] == valeur_104]
                        rubrique_valeur_5.loc[rubrique_valeur_5.index[0], 'Rubr'] = valeur_204
                        rubrique_valeur_5.loc[rubrique_valeur_5.index[1], 'Rubr'] = valeur_304
                        valeurs_de_a_remplir = list(self.rubriques_en_doublon_2_selectionnees.values())
                        sous_liste_valeurs_8 = valeurs_de_a_remplir[8]
                        sous_liste_valeurs_9 = valeurs_de_a_remplir[9]
                        rubrique_valeur_5.loc[rubrique_valeur_5.index[0], 'A Remplir'] = rubrique_valeur_5.loc[rubrique_valeur_5.index[0], sous_liste_valeurs_8]
                        rubrique_valeur_5.loc[rubrique_valeur_5.index[1], 'A Remplir'] = rubrique_valeur_5.loc[rubrique_valeur_5.index[0], sous_liste_valeurs_9]
                        print("Remplacement de la valeur de A remplir dans le dictionnaire 5 : ")
                        print(rubrique_valeur_5)
                        logging.info("Remplacement de la valeur de A remplir dans le dictionnaire 5 : ")
                        logging.info(rubrique_valeur_5)

                # S'assurer qu'il y a des valeurs dans le dictionnaire 6   
                if nombre_de_dictionnaires >= 6:
                    if valeurs[5]:                    
                        sous_liste_valeurs_105 = valeurs[5]
                        valeur_105, valeur_205, valeur_305 = sous_liste_valeurs_105
                        rubrique_valeur_6 = resultat_fusion_mat_rub_doublons_null.loc[resultat_fusion_mat_rub_doublons_null['Rubr'] == valeur_105]
                        rubrique_valeur_6.loc[rubrique_valeur_6.index[0], 'Rubr'] = valeur_205
                        rubrique_valeur_6.loc[rubrique_valeur_6.index[1], 'Rubr'] = valeur_305
                        valeurs_de_a_remplir = list(self.rubriques_en_doublon_2_selectionnees.values())
                        sous_liste_valeurs_10 = valeurs_de_a_remplir[10]
                        sous_liste_valeurs_11 = valeurs_de_a_remplir[11]
                        rubrique_valeur_6.loc[rubrique_valeur_6.index[0], 'A Remplir'] = rubrique_valeur_6.loc[rubrique_valeur_6.index[0], sous_liste_valeurs_10]
                        rubrique_valeur_6.loc[rubrique_valeur_6.index[1], 'A Remplir'] = rubrique_valeur_6.loc[rubrique_valeur_6.index[0], sous_liste_valeurs_11]
                        print("Remplacement de la valeur de A remplir dans le dictionnaire 6 : ")
                        print(rubrique_valeur_6)

                if nombre_de_dictionnaires >= 7:
                    if valeurs[6]:                    
                        sous_liste_valeurs_106 = valeurs[6]
                        valeur_106, valeur_206, valeur_306 = sous_liste_valeurs_106
                        rubrique_valeur_7 = resultat_fusion_mat_rub_doublons_null.loc[resultat_fusion_mat_rub_doublons_null['Rubr'] == valeur_106]
                        rubrique_valeur_7.loc[rubrique_valeur_7.index[0], 'Rubr'] = valeur_206
                        rubrique_valeur_7.loc[rubrique_valeur_7.index[1], 'Rubr'] = valeur_306
                        valeurs_de_a_remplir = list(self.rubriques_en_doublon_2_selectionnees.values())
                        sous_liste_valeurs_12 = valeurs_de_a_remplir[12]
                        sous_liste_valeurs_13 = valeurs_de_a_remplir[13]
                        rubrique_valeur_7.loc[rubrique_valeur_7.index[0], 'A Remplir'] = rubrique_valeur_7.loc[rubrique_valeur_7.index[0], sous_liste_valeurs_12]
                        rubrique_valeur_7.loc[rubrique_valeur_7.index[1], 'A Remplir'] = rubrique_valeur_7.loc[rubrique_valeur_7.index[0], sous_liste_valeurs_13]
                        print("Remplacement de la valeur de A remplir dans le dictionnaire 7 : ")
                        print(rubrique_valeur_7)
                        logging.info("Remplacement de la valeur de A remplir dans le dictionnaire 7 : ")
                        logging.info(rubrique_valeur_7)

                if nombre_de_dictionnaires >= 8:
                    if valeurs[7]:                    
                        sous_liste_valeurs_107 = valeurs[7]
                        valeur_107, valeur_207, valeur_307 = sous_liste_valeurs_107
                        rubrique_valeur_8 = resultat_fusion_mat_rub_doublons_null.loc[resultat_fusion_mat_rub_doublons_null['Rubr'] == valeur_107]
                        rubrique_valeur_8.loc[rubrique_valeur_8.index[0], 'Rubr'] = valeur_207
                        rubrique_valeur_8.loc[rubrique_valeur_8.index[1], 'Rubr'] = valeur_307
                        valeurs_de_a_remplir = list(self.rubriques_en_doublon_2_selectionnees.values())
                        sous_liste_valeurs_14 = valeurs_de_a_remplir[14]
                        sous_liste_valeurs_15 = valeurs_de_a_remplir[15]
                        rubrique_valeur_8.loc[rubrique_valeur_8.index[0], 'A Remplir'] = rubrique_valeur_8.loc[rubrique_valeur_8.index[0], sous_liste_valeurs_14]
                        rubrique_valeur_8.loc[rubrique_valeur_8.index[1], 'A Remplir'] = rubrique_valeur_8.loc[rubrique_valeur_8.index[0], sous_liste_valeurs_15]
                        print("Remplacement de la valeur de A remplir dans le dictionnaire 8 : ")
                        print(rubrique_valeur_8)
                        logging.info("Remplacement de la valeur de A remplir dans le dictionnaire 8 : ")
                        logging.info(rubrique_valeur_8)
                
                if nombre_de_dictionnaires == 1:
                    # Fusionner les DataFrames sur la colonne 'Rubr'
                    rubrique_valeur_final = rubrique_valeur_1
                    print("rubrique_valeur_final")
                    print(rubrique_valeur_final)
                    logging.info("rubrique_valeur_final")
                    logging.info(rubrique_valeur_final)
                
                if nombre_de_dictionnaires == 2:
                    # Fusionner les DataFrames sur la colonne 'Rubr'
                    rubrique_valeur_final = pd.concat([rubrique_valeur_1, rubrique_valeur_2], ignore_index=True)
                    print("rubrique_valeur_final")
                    print(rubrique_valeur_final)
                    logging.info("rubrique_valeur_final")
                    logging.info(rubrique_valeur_final)

                if nombre_de_dictionnaires == 3:
                    # Fusionner les DataFrames sur la colonne 'Rubr'
                    rubrique_valeur_final = pd.concat([rubrique_valeur_1, rubrique_valeur_2, rubrique_valeur_3], ignore_index=True)
                    print("rubrique_valeur_final")
                    print(rubrique_valeur_final)
                    logging.info("rubrique_valeur_final")
                    logging.info(rubrique_valeur_final)

                if nombre_de_dictionnaires == 4:
                    # Fusionner les DataFrames sur la colonne 'Rubr'
                    rubrique_valeur_final = pd.concat([rubrique_valeur_1, rubrique_valeur_2, rubrique_valeur_3, rubrique_valeur_4], ignore_index=True)
                    print("rubrique_valeur_final")
                    print(rubrique_valeur_final)
                    logging.info("rubrique_valeur_final")
                    logging.info(rubrique_valeur_final)

                if nombre_de_dictionnaires == 5:
                    # Fusionner les DataFrames sur la colonne 'Rubr'
                    rubrique_valeur_final = pd.concat([rubrique_valeur_1, rubrique_valeur_2, rubrique_valeur_3, rubrique_valeur_4, rubrique_valeur_5], ignore_index=True)
                    print("rubrique_valeur_final")
                    print(rubrique_valeur_final)
                    logging.info("rubrique_valeur_final")
                    logging.info(rubrique_valeur_final)

                if nombre_de_dictionnaires == 6:
                    # Fusionner les DataFrames sur la colonne 'Rubr'
                    rubrique_valeur_final = pd.concat([rubrique_valeur_1, rubrique_valeur_2, rubrique_valeur_3, rubrique_valeur_4, rubrique_valeur_5, rubrique_valeur_6], ignore_index=True)
                    print("rubrique_valeur_final")
                    print(rubrique_valeur_final)
                    logging.info("rubrique_valeur_final")
                    logging.info(rubrique_valeur_final)

                if nombre_de_dictionnaires == 7:
                    # Fusionner les DataFrames sur la colonne 'Rubr'
                    rubrique_valeur_final = pd.concat([rubrique_valeur_1, rubrique_valeur_2, rubrique_valeur_3, rubrique_valeur_4, rubrique_valeur_5, rubrique_valeur_6, rubrique_valeur_7], ignore_index=True)
                    print("rubrique_valeur_final")
                    print(rubrique_valeur_final)
                    logging.info("rubrique_valeur_final")
                    logging.info(rubrique_valeur_final)

                if nombre_de_dictionnaires == 8:
                    # Fusionner les DataFrames sur la colonne 'Rubr'
                    rubrique_valeur_final = pd.concat([rubrique_valeur_1, rubrique_valeur_2, rubrique_valeur_3, rubrique_valeur_4, rubrique_valeur_5, rubrique_valeur_6, rubrique_valeur_7, rubrique_valeur_8], ignore_index=True)
                    print("rubrique_valeur_final")
                    print(rubrique_valeur_final)
                    logging.info("rubrique_valeur_final")
                    logging.info(rubrique_valeur_final)


        else:
            print("Aucune rubrique en doubles n'a été trouvée pour l'agence sélectionnée.")
            logging.info("Aucune rubrique en doubles n'a été trouvée pour l'agence sélectionnée.")



        # Ajouter une valeur à toutes les lignes de la colonne 'Rubr' dans rubriques_en_doublon_LE NOM DE L'AGENCE : 
        #for rubrique, valeurs in rubriques_en_doublon_ASNIERES.items():
        #for rubrique, valeurs in self.rubriques_en_doublon_selectionnees.items():
        #    # Séparer les valeurs en deux listes distinctes
        #    valeur_1, valeur_2, valeur_3, valeur_4 = valeurs
        #    for index, row in resultat_fusion_mat_rub_doublons_null.iterrows():
        #        if row['Rubr'] == rubrique:
        #            # Attribuer les valeurs à la colonne 'A Remplir' en fonction de la parité
        #            if index % 2 == 0:
        #                resultat_fusion_mat_rub_doublons_null.loc[index, 'A Remplir'] = resultat_fusion_mat_rub_doublons_null.loc[index, valeur_1] # valeur_1
        #                resultat_fusion_mat_rub_doublons_null.loc[index, 'Rubr'] += ' ' + valeur_3
        #            else:
        #                resultat_fusion_mat_rub_doublons_null.loc[index, 'A Remplir'] = resultat_fusion_mat_rub_doublons_null.loc[index, valeur_2] # valeur_2
        #                resultat_fusion_mat_rub_doublons_null.loc[index, 'Rubr'] += ' ' + valeur_4

        
        # Debug
        # Maintenant, vous avez la liste de lignes filtrées prêtes à être utilisées pour écrire dans le fichier final
        
        #print("Lignes filtrées pour les doublons :")
        #print(resultat_fusion_mat_rub_doublons_null)
        

        ###########################################
        #   Cas specifique des lignes en triplon  # 
        ###########################################       

        logging.info("###########################################")
        logging.info("#   Cas specifique des lignes en triplon  #")
        logging.info("###########################################")

        # Compter le nombre d'occurrences de chaque valeur dans la colonne 'Rubr'
        rubrique_counts = resultat_fusion_mat_rub_doublons_null['Rubr'].value_counts()

        # Filtrer les lignes qui apparaissent trois fois avec la même valeur dans la colonne 'Rubr'
        rubriques_triples = rubrique_counts[rubrique_counts == 3].index

        # Filtrer les lignes correspondantes
        resultat_fusion_mat_rub_triples = resultat_fusion_mat_rub_doublons_null[resultat_fusion_mat_rub_doublons_null['Rubr'].isin(rubriques_triples)]

        # Afficher les lignes filtrées pour les triples occurrences
        print("Lignes filtrées pour les triplons :")
        print(resultat_fusion_mat_rub_triples)

        logging.info("Lignes filtrées pour les triplons :")
        logging.info(resultat_fusion_mat_rub_triples)

        # Vérifier si self.rubriques_en_triplon_selectionnees est défini
        if self.rubriques_en_triplon_selectionnees:
            # Récupérer les valeurs du dictionnaire self.rubriques_en_triplon_selectionnees
            valeurs = list(self.rubriques_en_triplon_selectionnees.values())

            # S'assurer qu'il y a des valeurs dans le dictionnaire
            if valeurs:
                # Accéder à la première (et seule) sous-liste de valeurs
                sous_liste_valeurs = valeurs[0]  

                # Extraire les valeurs individuelles dans des variables distinctes
                valeur_1, valeur_2, valeur_3 = sous_liste_valeurs

                # Afficher les valeurs individuelles
                print("Valeur 1:", valeur_1)
                print("Valeur 2:", valeur_2)
                print("Valeur 3:", valeur_3)
                logging.info("Valeur 1:", valeur_1)
                logging.info("Valeur 2:", valeur_2)
                logging.info("Valeur 3:", valeur_3)

        else:
            print("Aucune rubrique en triple n'a été trouvée pour l'agence sélectionnée.")
            logging.info("Aucune rubrique en triple n'a été trouvée pour l'agence sélectionnée.")

        # Accéder à la première ligne de resultat_fusion_mat_rub_triples
        # Remplace sur chaque ligne, la valeur de 'Rubr' (4 digits) par la liste dipo dans gestion_des_agences.py rubriques_en_triplon_*, plus particulierement : '5110': ['5110 base', '5110 salarié montant', '5110 à retenir'],
        if self.rubriques_en_triplon_selectionnees:
            if not resultat_fusion_mat_rub_triples.empty:
                resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[0], 'Rubr'] = valeur_1
                resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[1], 'Rubr'] = valeur_2
                resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[2], 'Rubr'] = valeur_3

                # Afficher le résultat après la modification
                print("Résultat après remplacement :")
                print(resultat_fusion_mat_rub_triples)
                logging.info("Résultat après remplacement :")
                logging.info(resultat_fusion_mat_rub_triples)
        else:
            print("Aucune rubrique en triple n'a été trouvée pour l'agence sélectionnée.")
            logging.info("Aucune rubrique en triple n'a été trouvée pour l'agence sélectionnée.")

        # TEST = 
        # Vérifier si self.rubriques_en_triplon_2_selectionnees est défini
        if self.rubriques_en_triplon_2_selectionnees:
            # Récupérer les valeurs du dictionnaire self.rubriques_en_triplon_2_selectionnees
            valeurs_triplon_2 = list(self.rubriques_en_triplon_2_selectionnees.values())

            # S'assurer qu'il y a des valeurs dans le dictionnaire
            if valeurs_triplon_2:
                print("valeurs_triplon_2:")
                print(valeurs_triplon_2)
                logging.info("valeurs_triplon_2:")
                logging.info(valeurs_triplon_2)

                # Extraire les valeurs individuelles dans des variables distinctes
                valeur_1_triplon_2, valeur_2_triplon_2, valeur_3_triplon_2 = valeurs_triplon_2

                # Afficher les valeurs individuelles
                print("Valeur 1 triplon 2:", valeur_1_triplon_2)
                print("Valeur 2 triplon 2:", valeur_2_triplon_2)
                print("Valeur 3 triplon 2:", valeur_3_triplon_2)
                logging.info("Valeur 1 triplon 2:", valeur_1_triplon_2)
                logging.info("Valeur 2 triplon 2:", valeur_2_triplon_2)
                logging.info("Valeur 3 triplon 2:", valeur_3_triplon_2)

        else:
            print("Aucune rubrique en triplon 2 n'a été trouvée pour l'agence sélectionnée.")
            logging.info("Aucune rubrique en triplon 2 n'a été trouvée pour l'agence sélectionnée.")

        # Remplacer sur chaque ligne la valeur de 'Rubr' par les valeurs du triplon 2
        if self.rubriques_en_triplon_2_selectionnees:
            if not resultat_fusion_mat_rub_triples.empty:
                resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[0], 'A Remplir'] = resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[0], valeur_1_triplon_2] 
                resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[1], 'A Remplir'] = resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[1], valeur_2_triplon_2]
                resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[2], 'A Remplir'] = resultat_fusion_mat_rub_triples.loc[resultat_fusion_mat_rub_triples.index[2], valeur_3_triplon_2]

                # Afficher le résultat après le remplacement
                # resultat_fusion_mat_rub_triples['A Remplir'] = resultat_fusion_mat_rub_triples['A Remplir'].abs()
                print("Résultat après remplacement triplon 2 :")
                print(resultat_fusion_mat_rub_triples)
                logging.info("Résultat après remplacement triplon 2 :")
                logging.info(resultat_fusion_mat_rub_triples)
        else:
            print("Aucune rubrique en triple n'a été trouvée pour l'agence sélectionnée.")
            logging.info("Aucune rubrique en triple n'a été trouvée pour l'agence sélectionnée.")

        #############################################
        # Ecriture des donnes dans le fichier final #
        #############################################
        #
        logging.info("#############################################")
        logging.info("# Ecriture des donnes dans le fichier final #")
        logging.info("#############################################")

        # Charger le fichier Excel
        excel_file_3 = self.fichier()
        print("ok debug 1")
        print(excel_file_3)
        excel_file_3_path = self.fichier()
        print("ok debug 2")
        print(excel_file_3_path)
        if excel_file_3 is not None:
            print("ok debug 3")
            wb = load_workbook(excel_file_3)
            print("ok debug 4")
            # Accéder à la feuille "Tempo-Banco"
            ws = wb["Tempo-Banco"]
            print("ok debug 5")
            print(ws)
            # Initialisation des variables pour les lignes 'Journal matricule' et 'Journal de rubriques'
            ligne_Jmatricule = None
            ligne_Jrubriques = None


            # Parcourir les lignes pour trouver les lignes 'Journal matricule' et 'Journal de rubriques'
            for row in ws.iter_rows():
                for cell in row:
                    if self.selected_text == "ARRAS":
                        if cell.value == 'Journal de matricule':
                            ligne_Jmatricule = cell.row
                        elif cell.value == 'Journal de rubriques':
                            ligne_Jrubriques = cell.row
                        elif cell.value == 'Agence patronal montant':
                            ligne_Agence_patronal_montant = cell.row
                        
                    #if self.selected_text == "MAUBEUGE":
                    #    if cell.value == 'Journal de matricule':
                    #        ligne_Jmatricule = cell.row
                    #    elif cell.value == 'Journal de rubriques':
                    #        ligne_Jrubriques = cell.row
                    #    elif cell.value == 'Agence patronal montant':
                    #        ligne_Agence_patronal_montant = cell.row

                    else :
                        if cell.value == 'Journal matricule':
                            ligne_Jmatricule = cell.row
                        elif cell.value == 'Journal de rubriques':
                            ligne_Jrubriques = cell.row
                        elif cell.value == 'Agence patronal montant':
                            ligne_Agence_patronal_montant = cell.row


            if ligne_Jrubriques:
                # Écrire les valeurs de resultat_fusion_3 aux bonnes positions dans le fichier Excel
                print("=== DEBUG ÉCRITURE resultat_fusion_3 ===")
                for index, row in resultat_fusion_3.iterrows():
                    rubr_ou_titre = row['Rubr'] if pd.notnull(row['Rubr']) else row['Titre']
                    valeur_a_ecrire = row['A Remplir']
                    
                    print(f"Recherche de '{rubr_ou_titre}' pour écrire la valeur: {valeur_a_ecrire}")
                    
                    # Chercher la ligne correspondante dans le fichier Excel
                    ligne_trouvee = False
                    for row_idx in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row_idx, column=1).value
                        if cell_value and str(cell_value).strip() == str(rubr_ou_titre).strip():
                            ws.cell(row=row_idx, column=3, value=valeur_a_ecrire)
                            print(f"✅ Écrit '{valeur_a_ecrire}' à la ligne {row_idx} pour '{rubr_ou_titre}'")
                            ligne_trouvee = True
                            break
                    
                    if not ligne_trouvee:
                        print(f"❌ Ligne non trouvée pour '{rubr_ou_titre}'")
                
                print("=== FIN DEBUG ÉCRITURE ===")

            
            if ligne_Jmatricule:
                # Écrire les valeurs dans la colonne 3 après la ligne 'Journal matricule'
                ligne_ecriture = ligne_Jmatricule + 1
                colonne_ecriture = 3  # Colonnes sont 1-indexed
                for index, row in resultat_fusion_2.iterrows():
                    if pd.notnull(row['A Remplir']):  # Vérifier si la valeur n'est pas nulle
                        ws.cell(row=ligne_ecriture, column=colonne_ecriture, value=row['A Remplir'])
                        ligne_ecriture += 1

            # test
            if ligne_Agence_patronal_montant:
                ligne_ecriture = ligne_Agence_patronal_montant
                colonne_ecriture = 3  # Colonnes sont 1-indexed
                for index, row in resultat_fusion_4.iterrows():
                    if pd.notnull(row['A Remplir']):  # Vérifier si la valeur n'est pas nulle
                        ws.cell(row=ligne_ecriture, column=colonne_ecriture, value=row['A Remplir'])
                        ligne_ecriture += 1


            # les lignes qui commencent par 4 digits
            # Parcourir les lignes du DataFrame resultat_fusion_mat_rub pour écrire les valeurs dans les 4 digits
            for index, row in resultat_fusion_mat_rub.iterrows():
                rubr = str(row['Rubr'])[:4]  # Garder uniquement les 4 premiers caractères
                a_remplir = row['A Remplir']
                # a_remplir = abs(row['A Remplir'])  # Valeur absolue
                # Trouver l'index de la ligne correspondant à rubr dans la feuille "Tempo-Banco"
                for cell in ws["A"]:
                    # Comparer uniquement les 4 premiers caractères
                    if str(cell.value)[:4] == rubr:  
                        # Écrire la valeur de 'A Remplir' dans la colonne C de la même ligne
                        ws[f"C{cell.row}"] = a_remplir
                        break

            # gestion des doublons
            # Parcourir les lignes de resultat_fusion_mat_rub_doublons_null pour écrire les valeurs dans la colonne 'A Remplir'
            #for index, row in resultat_fusion_mat_rub_doublons_null.iterrows():
            for index, row in rubrique_valeur_final.iterrows():
                rubrique = row['Rubr']
                a_remplir = row['A Remplir']
                # Trouver l'index de la ligne correspondant à rubrique 
                for row_idx in range(1, ws.max_row + 1):  # Parcourir toutes les lignes de la feuille
                    if ws.cell(row=row_idx, column=1).value == rubrique:
                        # Écrire la valeur de 'A Remplir' dans la colonne C de la même ligne
                        ws.cell(row=row_idx, column=3, value=a_remplir)
                        break

            # gestion des triplons
            if self.rubriques_en_triplon_selectionnees:
                for index, row in resultat_fusion_mat_rub_triples.iterrows():
                    rubrique = row['Rubr']
                    a_remplir = row['A Remplir']
                    #a_remplir = abs(row['A Remplir'])  # Valeur absolue
                    # Trouver l'index de la ligne correspondant à rubrique 
                    for row_idx in range(1, ws.max_row + 1):  # Parcourir toutes les lignes de la feuille
                        if ws.cell(row=row_idx, column=1).value == rubrique:
                            # Écrire la valeur de 'A Remplir' dans la colonne C de la même ligne
                            ws.cell(row=row_idx, column=3, value=a_remplir)
                            break
            

            # Sauvegarder les modifications dans le fichier Excel
            wb.save(excel_file_3)
            print("la premiere ecriture du fichier a ete realise avec succes !")
            print(excel_file_3_path)
            logging.info("Sauvegarder les modifications dans le fichier Excel")
            logging.info("la premiere ecriture du fichier a ete realise avec succes !")
            logging.info(excel_file_3_path)

            # Dans cette partie, nous allons rescanner le fichier qui vient d'etre ecrit a la recherche 
            # de valeurs null dans la colonne 3 entre le ligne juste en dessous de 'Journal cotisations' et la derniere ligne de 'Journal de rubriques'
            if excel_file_3_path is not None:
                wb2 = load_workbook(excel_file_3_path)
                # Accéder à la feuille "Tempo-Banco"
                ws2 = wb2["Tempo-Banco"]
                 # Parcourir les lignes pour trouver les lignes 'Journal matricule' et 'Journal de rubriques'
                for row in ws.iter_rows():
                    for cell in row:
                        if self.selected_text == "ARRAS":
                            if cell.value == 'Journal cotisations':
                                ligne_Jcotisations2 = cell.row
                            elif cell.value == 'Journal de matricule':
                                ligne_Jmatricule2 = cell.row
                            elif cell.value == 'Journal de rubriques':
                                ligne_Jrubriques2 = cell.row
                            elif cell.value == 'Total à payer':
                                ligne_TotalAPayer2 = cell.row

                        #if self.selected_text == "MAUBEUGE":
                        #    if cell.value == 'Journal cotisations':
                        #        ligne_Jcotisations2 = cell.row
                        #    elif cell.value == 'Journal de matricule':
                        #        ligne_Jmatricule2 = cell.row
                        #    elif cell.value == 'Journal de rubriques':
                        #        ligne_Jrubriques2 = cell.row
                        #    elif cell.value == 'Total à payer':
                        #        ligne_TotalAPayer2 = cell.row

                        else :
                            if cell.value == 'Journal cotisations':
                                ligne_Jcotisations2 = cell.row
                            elif cell.value == 'Journal matricule':
                                ligne_Jmatricule2 = cell.row
                            elif cell.value == 'Journal de rubriques':
                                ligne_Jrubriques2 = cell.row
                            elif cell.value == 'Total à payer':
                                ligne_TotalAPayer2 = cell.row
                
                if ligne_Jcotisations2:
                    print("ligne_Jcotisations2")
                    # Parcourir les lignes entre 'ligne_Jcotisations2' et 'ligne_Jmatricule2'
                    for row_idx in range(ligne_Jcotisations2 + 1, ligne_Jmatricule2):
                        if ws2.cell(row=row_idx, column=1).value is not None and not ws2.cell(row=row_idx, column=3).value:
                            # Écrire la valeur 0 dans la colonne C si la colonne A contient une valeur non null et la colonne C est vide
                            ws2.cell(row=row_idx, column=3, value=0)
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=3).value == '     ':
                            # Si la valeur de la colonne C est '     ', alors elle vaut 0
                            ws2.cell(row=row_idx, column=3, value=0)
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=1).value.startswith('3202 base'):
                            # Exception pour '3202 base': ne pas convertir les valeurs négatives en valeurs absolues
                            pass
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=1).value.startswith('3402 base'):
                            # Exception pour '3402 base': ne pas convertir les valeurs négatives en valeurs absolues
                            pass
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=1).value.startswith('3404 base'):
                            # Exception pour '3404 base': ne pas convertir les valeurs négatives en valeurs absolues
                            pass
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=1).value.startswith('3006 base'):
                            # Exception pour '3006 base': ne pas convertir les valeurs négatives en valeurs absolues
                            pass
                        elif ws2.cell(row=row_idx, column=3).value is not None and float(ws2.cell(row=row_idx, column=3).value) < 0:
                            # Remplacer les valeurs négatives par leur valeur absolue
                            ws2.cell(row=row_idx, column=3, value=abs(float(ws2.cell(row=row_idx, column=3).value)))


                if ligne_Jmatricule2:
                    print("ligne_Jmatricule2")
                    for row_idx in range(ligne_Jmatricule2 + 1, ligne_Jrubriques2):
                        if ws2.cell(row=row_idx, column=1).value is not None and not ws2.cell(row=row_idx, column=3).value:
                            # Écrire la valeur 0 dans la colonne C si la colonne A contient une valeur non null et la colonne C est vide
                            ws2.cell(row=row_idx, column=3, value=0)
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=3).value == '     ': # Asniere
                            # Si la valeur de la colonne C est '     ', alors elle vaut 0
                            ws2.cell(row=row_idx, column=3, value=0)
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=1).value.startswith('3006 base'):
                            # Exception pour '3006 base': ne pas convertir les valeurs négatives en valeurs absolues
                            pass
                
                if ligne_Jrubriques2:
                    print("ligne_Jrubriques2")
                    for row_idx in range(ligne_Jrubriques2 + 1, ligne_TotalAPayer2):
                        if ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=3).value is None:
                            # Écrire la valeur 0 dans la colonne C si la colonne A contient une valeur non null et la colonne C est vide
                            ws2.cell(row=row_idx, column=3, value=0)
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=3).value == '      ': # Lens
                            # Si la valeur de la colonne C est '     ', alors elle vaut 0
                            ws2.cell(row=row_idx, column=3, value=0)
                        elif ws2.cell(row=row_idx, column=1).value is not None and ws2.cell(row=row_idx, column=1).value.startswith('3006 base'):
                            # Exception pour '3006 base': ne pas convertir les valeurs négatives en valeurs absolues
                            pass
                        elif ws2.cell(row=row_idx, column=3).value is not None and float(ws2.cell(row=row_idx, column=3).value) < 0:
                            # Remplacer les valeurs négatives par leur valeur absolue
                            ws2.cell(row=row_idx, column=3, value=abs(float(ws2.cell(row=row_idx, column=3).value)))
                                
                # Sauvegarder les modifications dans le fichier Excel
                wb2.save(excel_file_3_path)
                print("la seconde ecriture du fichier a ete realise avec succes !")
                print(excel_file_3_path)
                logging.info("la seconde ecriture du fichier a ete realise avec succes !")
                logging.info(excel_file_3_path)
            
            # Cache le bouton pendant l'execution afin d'empecher 2 lancements simultanes :
            self.dialog_ui.pushButton_3.setEnabled(True)

            # PopUp de succès :           
            self.show_success_dialog(f"Fichier {excel_file_3_path} écrit avec succès !")
            print(excel_file_3_path)

            

    ######################################
    # Fonctions popup Success et Failed  #
    ######################################
    
    def show_success_dialog(self, message):
        show_success_dialog = SuccessDialog(message)
        show_success_dialog.exec_()

    def show_error_dialog(self, error_message):
        error_dialog = ErrorDialog(error_message)
        error_dialog.exec_()
    
    def show_error_running_dialog(self, error_message):
        error_dialog = ErrorDialogAlreadyRunning(error_message)
        error_dialog.exec_()

    #############################################################
    # fonctions des boutons du fichier gestion_paie_ui_dialog : #
    #############################################################

    # Quitter l'application :
    def Quitter(self):
         # Supprimer le fichier 'pid_projet_janus.tmp' s'il existe
        filename = "pid_projet_janus.tmp"
        if os.path.exists(filename):
            os.remove(filename)
        
        # Quitter l'application

        QtWidgets.QApplication.quit()  
    
    # Gestion des agences :
    def GestionDesAgences(self):

        # Déterminer le chemin absolu du fichier gestion_des_agences.py
        script_dir = os.path.dirname(os.path.abspath(__file__))
        gestion_agences_path = os.path.join(script_dir, 'gestion_des_agences.py')
        os.system(f'start notepad "{gestion_agences_path}"')  # Ouvre le fichier avec Notepad sur Windows
    
    # Fichiers en input : 
    def Cotisation_file(self, dialog_ui):
        options = QFileDialog.Options()
        fileName_Cotisation, _ = QFileDialog.getOpenFileName(self,"Choisissez le fichier Excel Cotisation", "","Excel Files (*.xlsx)", options=options)
        if fileName_Cotisation:

            try:
                cotisation = pd.read_excel(fileName_Cotisation, engine='openpyxl')
                # Liste des noms de colonnes attendus
                cotisation_columns = ['Rubr', 'Patronal Mont', 'Salarié Mont']
                # Vérification des colonnes
                if not all(column in cotisation.columns for column in cotisation_columns):
                    raise logging.info("Une erreur s'est produite lors de la lecture du fichier Cotisation.")
                # Rendre visible le logo IF_OK_excel
                dialog_ui.IF_OK_excel.setVisible(True)
                self.fileName_Cotisation = fileName_Cotisation
            except Exception as e:
                print("Une erreur s'est produite lors de la lecture du fichier Cotisation :", str(e))
                logging.info("Une erreur s'est produite lors de la lecture du fichier Cotisation :")
                self.show_error_dialog("Une erreur s'est produite lors de la lecture du fichier JALCOT")


    def Matricule_file(self, dialog_ui):
        options = QFileDialog.Options()
        fileName_Matricule, _ = QFileDialog.getOpenFileName(self,"Choisissez le fichier Excel Matricule", "","Excel Files (*.xlsx)", options=options)
        if fileName_Matricule:
            try:
                matricule = pd.read_excel(fileName_Matricule, engine='openpyxl')
                matricule_columns = ['Nom', 'Tranche A', 'Tranche B']
                if not all(column in matricule.columns for column in matricule_columns):
                    raise ValueError("Une erreur s'est produite lors de la lecture du fichier Matricule.")
                # Rendre visible IF_OK_Excel
                dialog_ui.IF_OK_excel_2.setVisible(True)
                self.fileName_Matricule = fileName_Matricule
            except Exception as e:
                print("Une erreur s'est produite lors de la lecture du fichier Matricule :", str(e))
                logging.info("Une erreur s'est produite lors de la lecture du fichier Matricule")
                self.show_error_dialog("Une erreur s'est produite lors de la lecture du fichier Matriule")

    def Rubrique_file(self, dialog_ui):
        options = QFileDialog.Options()
        fileName_Rubrique, _ = QFileDialog.getOpenFileName(self,"Choisissez le fichier Excel Rubrique", "","Excel Files (*.xlsx)", options=options)
        if fileName_Rubrique:
            try:
                rubrique = pd.read_excel(fileName_Rubrique, engine='openpyxl')
                rubrique_columns = ['Rubrique']
                # Vérification des colonnes
                if not all(column in rubrique.columns for column in rubrique_columns):
                    raise ValueError("Une erreur s'est produite lors de la lecture du fichier Rubrique.")
                # Rendre visible IF_OK_Excel
                dialog_ui.IF_OK_excel_3.setVisible(True)
                self.fileName_Rubrique = fileName_Rubrique
            except Exception as e:
                print("Une erreur s'est produite lors de la lecture du fichier Rubrique :", str(e))
                logging.info("Une erreur s'est produite lors de la lecture du fichier Rubrique")
                self.show_error_dialog("Une erreur s'est produite lors de la lecture du fichier JALRUB")


    def Sortie_file(self, dialog_ui):
        options = QFileDialog.Options()
        fileName_Sortie, _ = QFileDialog.getOpenFileName(self,"Choisissez le fichier Excel à écrire", "","Excel Files (*.xlsx)", options=options)
        if fileName_Sortie:
            # Rendre visible IF_OK_Excel
            dialog_ui.IF_OK_excel_4.setVisible(True)
            self.fileName_Sortie = fileName_Sortie

    # Définir la fonction appelée lorsqu'un nouvel élément est sélectionné dans la combobox pour le choix des agences
    logging.info("Définir la fonction appelée lorsqu'un nouvel élément est sélectionné dans la combobox pour le choix des agences")
    def on_index_changed(self, index, dialog_ui):
        # Récupérez le texte sélectionné dans la ComboBox de dialog_ui
        selected_text = dialog_ui.comboBox.currentText()
        print("Selected Text:", selected_text)  # Utilisez la valeur sélectionnée pour sélectionner le bon dictionnaire de correspondances
        logging.info("Selected Text:")
        logging.info(selected_text)
        self.selected_text = selected_text
        # Utilisez le texte sélectionné pour récupérer le bon dictionnaire de correspondances
        self.correspondances_selectionnees = CORRESPONDANCES_AGENCES.get(selected_text)
        if self.correspondances_selectionnees:
            # Utilise ce dictionnaire pour effectuer les remplacements dans la boucle plus haut
            print("self.correspondances_selectionnees")
            print(self.correspondances_selectionnees)
            logging.info("self.correspondances_selectionnees")
            logging.info(self.correspondances_selectionnees)

            #self.rubriques_en_doublon_selectionnees = getattr(gestion_des_agences, 'rubriques_en_doublon_' + selected_text, None)
            #if self.rubriques_en_doublon_selectionnees:
            #    # Utilisez rubriques_en_doublon_selectionnees pour effectuer les opérations souhaitées
            #    print("rubriques_en_doublon_selectionnees")
            #    print(self.rubriques_en_doublon_selectionnees)
            #else:
            #    print("Aucune rubrique en double n'a été trouvée pour l'agence sélectionnée.")
            
            self.rubriques_en_triplon_selectionnees = getattr(gestion_des_agences, 'rubriques_en_triplon_' + selected_text, None)
            if self.rubriques_en_triplon_selectionnees:
                # Utilisez rubriques_en_doublon_selectionnees pour effectuer les opérations souhaitées
                print("rubriques_en_triplon_selectionnees")
                print(self.rubriques_en_triplon_selectionnees)
                logging.info("rubriques_en_triplon_selectionnees")
                logging.info(self.rubriques_en_triplon_selectionnees)
            else:
                print("Aucune rubrique en triplons n'a été trouvée pour l'agence sélectionnée.")
                logging.info("Aucune rubrique en triplons n'a été trouvée pour l'agence sélectionnée.")
            
            self.rubriques_en_triplon_2_selectionnees = getattr(gestion_des_agences, 'rubriques_en_triplon_2_' + selected_text, None)
            if self.rubriques_en_triplon_2_selectionnees:
                print("rubriques_en_triplon_2_selectionnees")
                print(self.rubriques_en_triplon_2_selectionnees)
                logging.info("rubriques_en_triplon_2_selectionnees")
                logging.info(self.rubriques_en_triplon_2_selectionnees)             
            else:
                print("Aucune rubrique en triplon 2 n'a été trouvée pour l'agence sélectionnée.")
                logging.info("Aucune rubrique en triplon 2 n'a été trouvée pour l'agence sélectionnée.")

            self.rubriques_en_doublon_selectionnees = getattr(gestion_des_agences, 'rubriques_en_doublon_' + selected_text, None)
            if self.rubriques_en_doublon_selectionnees:
                print("rubriques_en_doublon selectionnees")
                print(self.rubriques_en_doublon_selectionnees)
                logging.info("rubriques_en_doublon selectionnees")
                logging.info(self.rubriques_en_doublon_selectionnees)
            else:
                print("Aucune rubrique en doublon n'a été trouvée pour l'agence sélectionnée.")
                logging.info("Aucune rubrique en doublon n'a été trouvée pour l'agence sélectionnée.")
            
            self.rubriques_en_doublon_2_selectionnees = getattr(gestion_des_agences, 'rubriques_en_doublon_2_' + selected_text, None)
            if self.rubriques_en_doublon_2_selectionnees:
                print("rubriques_en_doublon_2 selectionnees")
                print(self.rubriques_en_doublon_2_selectionnees)
                logging.info("rubriques_en_doublon_2 selectionnees")
                logging.info(self.rubriques_en_doublon_2_selectionnees)
            else:
                print("Aucune rubrique en doublon 2 n'a été trouvée pour l'agence sélectionnée.")
                logging.info("Aucune rubrique en doublon 2 n'a été trouvée pour l'agence sélectionnée.")
        else:
            print("Aucune correspondance n'a été trouvée pour l'agence sélectionnée.")
            logging.info("Aucune correspondance n'a été trouvée pour l'agence sélectionnée.")
        

    #############################################################

def main():
    # Lancement de l'app:
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    logging.info("Lancement du programme")
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()



